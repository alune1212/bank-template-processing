"""批量合并目录功能测试。"""

from __future__ import annotations

import json
import logging
import os
from pathlib import Path

import openpyxl
import pytest

from bank_template_processing.main import (
    _calculate_stats,
    _needs_transformations,
    apply_transformations,
    main,
)
from bank_template_processing.merge_folder import (
    MergeFolderError,
    infer_month_param_from_values,
    parse_merge_filename,
    prepare_merge_tasks,
    resolve_rule_group_for_template,
)


def _create_template_file(path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["姓名", "实发工资", "用途"])
    workbook.save(path)


def _create_generated_file(path: Path, rows: list[tuple[str, float, str]]) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["姓名", "实发工资", "用途"])
    for row in rows:
        sheet.append(list(row))
    workbook.save(path)


def _build_test_config(default_template: Path, crossbank_template: Path, month_type_enabled: bool = False) -> dict:
    return {
        "version": "2.0",
        "organization_units": {
            "苏州悦鸣服务外包有限公司": {
                "template_selector": {
                    "enabled": True,
                    "default_bank": "农业银行",
                    "default_group_name": "农行跨行",
                    "special_group_name": "招行跨行",
                },
                "default": {
                    "template_path": str(default_template),
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {
                        "姓名": {
                            "source_column": "姓名",
                            "target_column": "姓名",
                            "transform": "none",
                        },
                        "金额": {
                            "source_column": "实发工资",
                            "target_column": "实发工资",
                            "transform": "amount_decimal",
                        },
                    },
                    "transformations": {
                        "amount_decimal": {
                            "decimal_places": 2,
                            "rounding": "round",
                        }
                    },
                    "validation_rules": {},
                    "month_type_mapping": {
                        "enabled": month_type_enabled,
                        "target_column": "用途",
                        "month_format": "{month}月收入",
                        "bonus_value": "年终奖",
                        "compensation_value": "补偿金",
                    },
                },
                "crossbank": {
                    "template_path": str(crossbank_template),
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {
                        "姓名": {
                            "source_column": "姓名",
                            "target_column": "姓名",
                            "transform": "none",
                        },
                        "金额": {
                            "source_column": "实发工资",
                            "target_column": "实发工资",
                            "transform": "amount_decimal",
                        },
                    },
                    "transformations": {
                        "amount_decimal": {
                            "decimal_places": 2,
                            "rounding": "round",
                        }
                    },
                    "validation_rules": {},
                    "month_type_mapping": {
                        "enabled": month_type_enabled,
                        "target_column": "用途",
                        "month_format": "{month}月收入",
                        "bonus_value": "年终奖",
                        "compensation_value": "补偿金",
                    },
                },
            }
        },
    }


def test_parse_merge_filename_with_longest_unit_match():
    unit_names = ["苏州悦鸣", "苏州悦鸣服务外包有限公司"]
    file_path = Path("苏州悦鸣服务外包有限公司_农行跨行_4人_金额26616.76元.xlsx")

    parsed = parse_merge_filename(file_path, unit_names)

    assert parsed.unit_name == "苏州悦鸣服务外包有限公司"
    assert parsed.template_name == "农行跨行"
    assert parsed.count == 4
    assert parsed.amount == 26616.76


def test_parse_merge_filename_invalid_format_raises():
    unit_names = ["苏州悦鸣服务外包有限公司"]
    file_path = Path("苏州悦鸣服务外包有限公司_农行跨行.xlsx")

    with pytest.raises(MergeFolderError, match="文件名不符合合并命名规则"):
        parse_merge_filename(file_path, unit_names)


def test_resolve_rule_group_prefers_group_name_then_template_stem(tmp_path):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)
    config = _build_test_config(default_template, crossbank_template)

    rule_group_by_name, _ = resolve_rule_group_for_template(config, "苏州悦鸣服务外包有限公司", "农行跨行")
    assert rule_group_by_name == "default"

    rule_group_by_stem, _ = resolve_rule_group_for_template(
        config,
        "苏州悦鸣服务外包有限公司",
        "crossbank_template",
    )
    assert rule_group_by_stem == "crossbank"


def test_infer_month_param_from_values():
    month_type_mapping = {
        "enabled": True,
        "target_column": "用途",
        "month_format": "{month}月收入",
        "bonus_value": "年终奖",
        "compensation_value": "补偿金",
    }

    assert infer_month_param_from_values({"年终奖"}, month_type_mapping) == "年终奖"
    assert infer_month_param_from_values({"补偿金"}, month_type_mapping) == "补偿金"
    assert infer_month_param_from_values({"02月收入"}, month_type_mapping) == "02"

    with pytest.raises(MergeFolderError, match="同一分组存在冲突的月类型值"):
        infer_month_param_from_values({"年终奖", "补偿金"}, month_type_mapping)

    assert infer_month_param_from_values(
        {"01月收入", "年终奖"},
        month_type_mapping,
        allow_conflict=True,
    ) == "01"


def test_prepare_merge_tasks_stats_mismatch_raises(tmp_path):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)
    config = _build_test_config(default_template, crossbank_template)

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额200.00元.xlsx",
        [("张三", 100.0, "01月收入")],
    )

    with pytest.raises(MergeFolderError, match="人数校验失败"):
        prepare_merge_tasks(
            merge_folder_path=str(merge_dir),
            config=config,
            resolve_path_fn=lambda path: path,
            apply_transformations_fn=apply_transformations,
            needs_transformations_fn=_needs_transformations,
            calculate_stats_fn=_calculate_stats,
            needs_month_for_filename=False,
            logger=logging.getLogger(__name__),
        )


def test_prepare_merge_tasks_orders_by_mtime_not_filename(tmp_path):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)
    config = _build_test_config(default_template, crossbank_template)

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()

    filename_first = merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_1人_金额10.00元.xlsx"
    filename_later = merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_1人_金额20.00元.xlsx"

    _create_generated_file(filename_first, [("应后出现", 10.0, "01月收入")])
    _create_generated_file(filename_later, [("应先出现", 20.0, "01月收入")])

    older_ts = 1_700_000_000
    newer_ts = older_ts + 60
    os.utime(filename_later, (older_ts, older_ts))
    os.utime(filename_first, (newer_ts, newer_ts))

    tasks = prepare_merge_tasks(
        merge_folder_path=str(merge_dir),
        config=config,
        resolve_path_fn=lambda path: path,
        apply_transformations_fn=apply_transformations,
        needs_transformations_fn=_needs_transformations,
        calculate_stats_fn=_calculate_stats,
        needs_month_for_filename=False,
        logger=logging.getLogger(__name__),
    )

    assert len(tasks) == 1
    assert [row["姓名"] for row in tasks[0].group_data] == ["应先出现", "应后出现"]


def test_prepare_merge_tasks_mtime_tie_breaks_by_filename(tmp_path):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)
    config = _build_test_config(default_template, crossbank_template)

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()

    filename_10 = merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_1人_金额10.00元.xlsx"
    filename_20 = merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_1人_金额20.00元.xlsx"

    _create_generated_file(filename_10, [("文件名10在前", 10.0, "01月收入")])
    _create_generated_file(filename_20, [("文件名20在后", 20.0, "01月收入")])

    same_ts = 1_700_000_000
    os.utime(filename_10, (same_ts, same_ts))
    os.utime(filename_20, (same_ts, same_ts))

    tasks = prepare_merge_tasks(
        merge_folder_path=str(merge_dir),
        config=config,
        resolve_path_fn=lambda path: path,
        apply_transformations_fn=apply_transformations,
        needs_transformations_fn=_needs_transformations,
        calculate_stats_fn=_calculate_stats,
        needs_month_for_filename=False,
        logger=logging.getLogger(__name__),
    )

    assert len(tasks) == 1
    assert [row["姓名"] for row in tasks[0].group_data] == ["文件名10在前", "文件名20在后"]


def test_prepare_merge_tasks_skips_month_inference_when_filename_template_not_use_month(tmp_path, caplog):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)
    config = _build_test_config(default_template, crossbank_template, month_type_enabled=True)

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额300.00元.xlsx",
        [("张三", 100.0, "01月收入"), ("李四", 200.0, "年终奖")],
    )

    caplog.set_level(logging.INFO)
    tasks = prepare_merge_tasks(
        merge_folder_path=str(merge_dir),
        config=config,
        resolve_path_fn=lambda path: path,
        apply_transformations_fn=apply_transformations,
        needs_transformations_fn=_needs_transformations,
        calculate_stats_fn=_calculate_stats,
        needs_month_for_filename=False,
        logger=logging.getLogger(__name__),
    )

    assert len(tasks) == 1
    assert tasks[0].month_param == ""
    assert "检测到同一分组存在冲突的月类型值" not in caplog.text
    assert "跳过月份参数推断：输出文件名模板未使用 {month}" in caplog.text


def test_prepare_merge_tasks_infers_month_when_filename_template_uses_month(tmp_path, caplog):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)
    config = _build_test_config(default_template, crossbank_template, month_type_enabled=True)

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额300.00元.xlsx",
        [("张三", 100.0, "01月收入"), ("李四", 200.0, "年终奖")],
    )

    caplog.set_level(logging.INFO)
    tasks = prepare_merge_tasks(
        merge_folder_path=str(merge_dir),
        config=config,
        resolve_path_fn=lambda path: path,
        apply_transformations_fn=apply_transformations,
        needs_transformations_fn=_needs_transformations,
        calculate_stats_fn=_calculate_stats,
        needs_month_for_filename=True,
        logger=logging.getLogger(__name__),
    )

    assert len(tasks) == 1
    assert tasks[0].month_param == "01"
    assert "检测到同一分组存在冲突的月类型值" in caplog.text


def test_main_merge_folder_generates_result_files(tmp_path):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)

    config = _build_test_config(default_template, crossbank_template)
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()

    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额300.00元.xlsx",
        [("张三", 100.0, "01月收入"), ("李四", 200.0, "01月收入")],
    )
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_1人_金额50.00元.xlsx",
        [("王五", 50.0, "01月收入")],
    )
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_招行跨行_2人_金额100.00元.xlsx",
        [("赵六", 10.0, "01月收入"), ("孙七", 90.0, "01月收入")],
    )
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_招行跨行_1人_金额40.00元.xlsx",
        [("周八", 40.0, "01月收入")],
    )

    main(
        [
            "--merge-folder",
            str(merge_dir),
            "--config",
            str(config_path),
        ]
    )

    result_dir = merge_dir / "result"
    assert result_dir.exists()

    output_names = sorted(path.name for path in result_dir.glob("*.xlsx"))
    assert output_names == [
        "苏州悦鸣服务外包有限公司_农行跨行_3人_金额350.00元.xlsx",
        "苏州悦鸣服务外包有限公司_招行跨行_3人_金额140.00元.xlsx",
    ]


def test_main_merge_folder_keeps_row_month_values_for_mixed_months(tmp_path):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)

    config = _build_test_config(default_template, crossbank_template, month_type_enabled=True)
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()

    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额300.00元.xlsx",
        [("张三", 100.0, "01月收入"), ("李四", 200.0, "年终奖")],
    )

    main(
        [
            "--merge-folder",
            str(merge_dir),
            "--config",
            str(config_path),
        ]
    )

    output_path = merge_dir / "result" / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额300.00元.xlsx"
    assert output_path.exists()

    workbook = openpyxl.load_workbook(output_path)
    try:
        sheet = workbook.active
        assert sheet is not None
        assert sheet.cell(2, 3).value == "01月收入"
        assert sheet.cell(3, 3).value == "年终奖"
    finally:
        workbook.close()


def test_main_merge_folder_with_escaped_month_literal_template_skips_month_inference(tmp_path, caplog):
    default_template = tmp_path / "default_template.xlsx"
    crossbank_template = tmp_path / "crossbank_template.xlsx"
    _create_template_file(default_template)
    _create_template_file(crossbank_template)

    config = _build_test_config(default_template, crossbank_template, month_type_enabled=True)
    config_path = tmp_path / "config.json"
    config_path.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    merge_dir = tmp_path / "merge_input"
    merge_dir.mkdir()
    _create_generated_file(
        merge_dir / "苏州悦鸣服务外包有限公司_农行跨行_2人_金额300.00元.xlsx",
        [("张三", 100.0, "01月收入"), ("李四", 200.0, "年终奖")],
    )

    caplog.set_level(logging.INFO)
    main(
        [
            "--merge-folder",
            str(merge_dir),
            "--config",
            str(config_path),
            "--output-filename-template",
            "{unit_name}_{{month}}_{template_name}_{count}人_金额{amount:.2f}元{ext}",
        ]
    )

    assert "检测到同一分组存在冲突的月类型值" not in caplog.text

    output_path = merge_dir / "result" / "苏州悦鸣服务外包有限公司_{month}_农行跨行_2人_金额300.00元.xlsx"
    assert output_path.exists()
