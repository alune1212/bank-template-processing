"""pipeline 共享处理管线测试。"""

from __future__ import annotations

import logging

import openpyxl
import pytest

from bank_template_processing.pipeline import (
    ProcessingContext,
    build_reader,
    transform_rows,
    validate_rows,
    write_group_output,
)
from bank_template_processing.transformer import TransformError
from bank_template_processing.validator import ValidationError


def test_build_reader_falls_back_for_invalid_reader_options(caplog):
    caplog.set_level("WARNING")

    reader = build_reader({"reader_options": "bad"}, logger_instance=logging.getLogger(__name__))

    assert reader.header_row == 1
    assert "reader_options 配置无效" in caplog.text


def test_validate_rows_adds_context():
    with pytest.raises(ValidationError, match="数据校验失败（单位=单位A，规则组=default，第1条数据）"):
        validate_rows(
            [{"姓名": "张三"}],
            {"required_fields": ["金额"]},
            context=ProcessingContext(unit_name="单位A", rule_group="default"),
        )


def test_transform_rows_adds_context():
    with pytest.raises(TransformError, match="数据转换失败（单位=单位A，模板=模板A，第1条数据）"):
        transform_rows(
            [{"卡号": "123"}],
            {"card_number": {"luhn_validation": True}},
            {"卡号": {"source_column": "卡号", "transform": "card_number"}},
            context=ProcessingContext(unit_name="单位A", template_name="模板A"),
        )


def test_write_group_output_uses_shared_writer(tmp_path):
    template_path = tmp_path / "template.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["姓名", "金额"])
    workbook.save(template_path)

    output_path = tmp_path / "out.xlsx"
    write_group_output(
        group_data=[{"姓名": "张三", "金额": 100}],
        group_config={
            "field_mappings": {
                "姓名": {"source_column": "姓名", "target_column": "姓名"},
                "金额": {"source_column": "金额", "target_column": "金额", "transform": "amount_decimal"},
            },
            "header_row": 1,
            "start_row": 2,
            "transformations": {},
        },
        template_path=str(template_path),
        output_path=output_path,
        month_param="01",
        logger_instance=logging.getLogger(__name__),
    )

    result = openpyxl.load_workbook(output_path)
    sheet = result.active
    assert sheet is not None
    assert sheet.cell(2, 1).value == "张三"
    assert sheet.cell(2, 2).value == 100
    result.close()
