"""Integration test for the Chinese column name bug fix with actual xlsx save"""

import tempfile
from pathlib import Path

import openpyxl
import pytest

from bank_template_processing.excel_writer import ExcelWriter


def test_chinese_column_name_with_xlsx_save_integration():
    """
    Integration test: Reproduces the exact bug scenario
    where Chinese column names would cause ValueError on wb.save()
    """
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)

        # Create a template file with Chinese headers (like config.json "crossbank" template)
        template_path = tmpdir / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add header row as in config: header_row = 2
        ws.cell(1, 1, "说明文字")
        ws.cell(1, 2, "填表说明")

        # Header row with Chinese column names
        ws.cell(2, 1, "编号")
        ws.cell(2, 2, "收款方账号")
        ws.cell(2, 3, "收款方户名")
        ws.cell(2, 4, "金额")
        ws.cell(2, 5, "开户行支行名称")
        ws.cell(2, 6, "用途（附言）")

        # Add some old data that should be cleared
        ws.cell(3, 1, "1")
        ws.cell(3, 2, "1234567890123456")
        ws.cell(3, 3, "张三")
        ws.cell(3, 4, 1000.0)
        ws.cell(3, 5, "北京支行")
        ws.cell(3, 6, "01月收入")

        wb.save(template_path)

        # Prepare data
        data = [
            {
                "工资卡卡号": "6228481234567890123",
                "姓名": "李四",
                "实发工资": 5000.0,
                "开户银行": "中国工商银行",
            },
            {
                "工资卡卡号": "6228489876543210987",
                "姓名": "王五",
                "实发工资": 6000.0,
                "开户银行": "中国建设银行",
            },
        ]

        # Field mappings using Chinese column names (like config.json)
        field_mappings = {
            "收款方账号": {
                "source_column": "工资卡卡号",
                "target_column": "收款方账号",  # Chinese column name - THE BUG CASE
                "transform": "none",
            },
            "收款方户名": {
                "source_column": "姓名",
                "target_column": "收款方户名",  # Chinese column name - THE BUG CASE
                "transform": "none",
            },
            "金额": {
                "source_column": "实发工资",
                "target_column": "金额",  # Chinese column name - THE BUG CASE
                "transform": "none",
            },
            "开户行支行名称": {
                "source_column": "开户银行",
                "target_column": "开户行支行名称",  # Chinese column name - THE BUG CASE
                "transform": "none",
            },
        }

        # Fixed values with Chinese column names
        fixed_values = {
            "是否农行账户": "否",  # Chinese column name - THE BUG CASE
        }

        # Month type mapping with Chinese column name
        month_type_mapping = {
            "enabled": True,
            "target_column": "用途（附言）",  # Chinese column name - THE BUG CASE
        }

        # Auto number with Chinese column name
        auto_number = {
            "enabled": True,
            "column_name": "编号",  # Chinese column name - THE BUG CASE
            "start_from": 1,
        }

        output_path = tmpdir / "output.xlsx"

        # THIS IS THE KEY TEST: write_excel should succeed without ValueError
        # Before the fix, this would fail with:
        # ValueError: Invalid column index 373584997465
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=2,
            start_row=3,
            mapping_mode="column_name",
            fixed_values=fixed_values,
            auto_number=auto_number,
            month_type_mapping=month_type_mapping,
            month_param="1",
        )

        # Verify the output file was created successfully
        assert output_path.exists(), "Output file should be created"

        # Verify the data was written correctly
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active

        # Header should be preserved
        assert ws_result.cell(2, 1).value == "编号"
        assert ws_result.cell(2, 2).value == "收款方账号"
        assert ws_result.cell(2, 3).value == "收款方户名"
        assert ws_result.cell(2, 4).value == "金额"
        assert ws_result.cell(2, 5).value == "开户行支行名称"

        # Data row 1
        assert ws_result.cell(3, 1).value == 1  # Auto number
        assert ws_result.cell(3, 2).value == "6228481234567890123"
        assert ws_result.cell(3, 3).value == "李四"
        assert ws_result.cell(3, 4).value == 5000.0
        assert ws_result.cell(3, 5).value == "中国工商银行"
        assert ws_result.cell(3, 6).value is None

        # Data row 2
        assert ws_result.cell(4, 1).value == 2  # Auto number
        assert ws_result.cell(4, 2).value == "6228489876543210987"
        assert ws_result.cell(4, 3).value == "王五"
        assert ws_result.cell(4, 4).value == 6000.0
        assert ws_result.cell(4, 5).value == "中国建设银行"
        assert ws_result.cell(4, 6).value is None

        # Old data should be cleared
        assert ws_result.cell(5, 1).value is None

        wb_result.close()

        print("✓ Integration test passed: Chinese column names work correctly with xlsx save")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
