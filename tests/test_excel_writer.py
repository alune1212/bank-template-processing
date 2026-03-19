"""测试Excel写入器模块"""

import openpyxl
import pytest

from bank_template_processing.excel_writer import ExcelError, ExcelWriter
from tests.spreadsheet_factories import write_xls_rows, write_xlsx_rows


class TestExcelWriter:
    """ExcelWriter类的测试用例"""

    def test_write_xlsx_file(self, tmp_path):
        """测试写入.xlsx文件"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["说明文字"], ["姓名", "年龄", "金额"]])

        # 准备数据
        data = [
            {"姓名": "张三", "年龄": 25, "金额": 1000.0},
            {"姓名": "李四", "年龄": 30, "金额": 2000.0},
        ]

        # 字段映射
        field_mappings = {
            "姓名": {"source_column": "姓名"},
            "年龄": {"source_column": "年龄"},
            "金额": {"source_column": "金额"},
        }

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=2,
            start_row=3,
            mapping_mode="column_name",
        )

        # 验证输出文件存在
        assert output_path.exists()

        # 验证数据
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(1, 1).value == "说明文字"  # 保留说明文字
        assert ws_result.cell(2, 1).value == "姓名"  # 保留表头
        assert ws_result.cell(3, 1).value == "张三"  # 第一行数据
        assert ws_result.cell(3, 2).value == 25
        assert ws_result.cell(4, 1).value == "李四"  # 第二行数据

        wb_result.close()

    def test_amount_written_as_number(self, tmp_path):
        """测试金额字段以数字类型写入，而不是字符串"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "金额"]])

        data = [
            {"姓名": "张三", "金额": 1000.50},
            {"姓名": "李四", "金额": 2000.75},
        ]

        field_mappings = {
            "姓名": {
                "source_column": "姓名",
                "target_column": "姓名",
                "transform": "none",
            },
            "金额": {
                "source_column": "金额",
                "target_column": "金额",
                "transform": "amount_decimal",
                "required": True,
            },
        }

        output_path = tmp_path / "output.xlsx"
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )

        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active

        assert ws_result.cell(2, 1).value == "张三"
        assert isinstance(ws_result.cell(2, 2).value, (int, float)), "金额应该是数字类型"
        assert ws_result.cell(2, 2).value == 1000.50

        assert ws_result.cell(3, 1).value == "李四"
        assert isinstance(ws_result.cell(3, 2).value, (int, float)), "金额应该是数字类型"
        assert ws_result.cell(3, 2).value == 2000.75

        wb_result.close()

    def test_write_xls_file(self, tmp_path):
        """测试写入.xls文件"""
        template_path = write_xls_rows(tmp_path / "template.xls", [["说明文字"], ["姓名", "年龄"]])

        # 准备数据
        data = [{"姓名": "张三", "年龄": 25}, {"姓名": "李四", "年龄": 30}]

        # 字段映射
        field_mappings = {
            "姓名": {"source_column": "姓名"},
            "年龄": {"source_column": "年龄"},
        }

        # 输出路径
        output_path = tmp_path / "output.xls"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=2,
            start_row=3,
            mapping_mode="column_name",
        )

        # 验证输出文件存在
        assert output_path.exists()

    def test_write_agriculture_bank_xls_template_layout(self, tmp_path):
        """测试农业银行模板写入时保留标题和表头"""
        import xlrd

        template_path = write_xls_rows(
            tmp_path / "农业银行模板.xls",
            [
                ["中国农业银行代发工资（农行）上传文件", "", "", "", ""],
                ["编号", "收款方账号", "收款方户名", "金额", "备注（附言）"],
            ],
        )

        data = [
            {"姓名": "张三", "工资卡卡号": "6228480402564890018", "实发工资": 1234.56},
            {"姓名": "李四", "工资卡卡号": "6228480402564890026", "实发工资": 78.9},
        ]
        field_mappings = {
            "收款方户名": {"source_column": "姓名", "target_column": "收款方户名", "transform": "none"},
            "收款方账号": {"source_column": "工资卡卡号", "target_column": "收款方账号", "transform": "card_number"},
            "金额": {"source_column": "实发工资", "target_column": "金额", "transform": "amount_decimal"},
        }

        output_path = tmp_path / "农业银行模板_output.xls"
        ExcelWriter().write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=2,
            start_row=3,
            mapping_mode="column_name",
            auto_number={"enabled": True, "column_name": "编号", "start_from": 1},
            month_type_mapping={"enabled": True, "target_column": "备注（附言）", "month_format": "{month}月收入"},
            month_param="01",
            clear_rows={"start_row": 3, "end_row": 200},
        )

        workbook = xlrd.open_workbook(str(output_path))
        sheet = workbook.sheet_by_index(0)
        assert sheet.cell_value(0, 0) == "中国农业银行代发工资（农行）上传文件"
        assert sheet.row_values(1)[:5] == ["编号", "收款方账号", "收款方户名", "金额", "备注（附言）"]
        assert sheet.row_values(2)[:5] == [1.0, "6228480402564890018", "张三", 1234.56, "01月收入"]
        assert sheet.row_values(3)[:5] == [2.0, "6228480402564890026", "李四", 78.9, "01月收入"]

    def test_fixed_values(self, tmp_path):
        """测试固定值功能"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "部门"]])

        # 准备数据
        data = [{"姓名": "张三"}, {"姓名": "李四"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 固定值
        fixed_values = {"B": "财务部"}

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            fixed_values=fixed_values,
        )

    def test_month_format_custom(self, tmp_path):
        """测试 month_format 生效"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "用途"]])

        data = [{"姓名": "张三"}]
        field_mappings = {"姓名": {"source_column": "姓名"}}
        month_type_mapping = {
            "enabled": True,
            "target_column": "用途",
            "month_format": "{month}月工资",
            "bonus_value": "年终奖",
            "compensation_value": "补偿金",
        }

        output_path = tmp_path / "output.xlsx"
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            month_type_mapping=month_type_mapping,
            month_param="1",
        )

        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 2).value == "01月工资"
        wb_result.close()

    def test_clear_rows_preserve_tail(self, tmp_path):
        """测试 clear_rows 仅清理数据区并保留尾部"""
        template_path = write_xlsx_rows(
            tmp_path / "template.xlsx",
            [["姓名"], ["旧数据1"], ["旧数据2"], ["旧数据3"], ["合计"]],
        )

        data = [{"姓名": "张三"}, {"姓名": "李四"}]
        field_mappings = {"姓名": {"source_column": "姓名"}}

        output_path = tmp_path / "output.xlsx"
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            clear_rows={"start_row": 2, "end_row": 4},
        )

        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(3, 1).value == "李四"
        assert ws_result.cell(5, 1).value == "合计"
        wb_result.close()

    def test_clear_rows_insert_rows_when_data_exceeds(self, tmp_path):
        """测试 clear_rows 数据超过范围时插入行"""
        template_path = write_xlsx_rows(
            tmp_path / "template.xlsx",
            [["姓名"], ["旧数据1"], ["旧数据2"], ["合计"]],
        )

        data = [{"姓名": "张三"}, {"姓名": "李四"}, {"姓名": "王五"}]
        field_mappings = {"姓名": {"source_column": "姓名"}}

        output_path = tmp_path / "output.xlsx"
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            clear_rows={"start_row": 2, "end_row": 3},
        )

        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(3, 1).value == "李四"
        assert ws_result.cell(4, 1).value == "王五"
        assert ws_result.cell(5, 1).value == "合计"
        wb_result.close()

    def test_clear_rows_xls_insufficient_range(self, tmp_path):
        """测试 XLS clear_rows 范围不足时抛错"""
        from bank_template_processing.config_loader import ConfigError

        template_path = write_xls_rows(tmp_path / "template.xls", [["姓名"], ["旧数据1"]])

        data = [{"姓名": "张三"}, {"姓名": "李四"}]
        field_mappings = {"姓名": {"source_column": "姓名"}}

        writer = ExcelWriter()
        with pytest.raises(ConfigError):
            writer.write_excel(
                template_path=str(template_path),
                data=data,
                field_mappings=field_mappings,
                output_path=str(tmp_path / "output.xls"),
                header_row=1,
                start_row=2,
                mapping_mode="column_name",
                clear_rows={"start_row": 2, "end_row": 2},
            )

    def test_auto_number(self, tmp_path):
        """测试自动编号功能"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["序号", "姓名"]])

        # 准备数据
        data = [{"姓名": "张三"}, {"姓名": "李四"}, {"姓名": "王五"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 自动编号
        auto_number = {"enabled": True, "column": "A", "start_from": 1}

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            auto_number=auto_number,
        )

        # 验证自动编号
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == 1
        assert ws_result.cell(3, 1).value == 2
        assert ws_result.cell(4, 1).value == 3

        wb_result.close()

    def test_bank_branch_mapping(self, tmp_path):
        """测试银行支行映射功能（已废弃，现在使用 field_mappings 替代）"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "部门"]])

        # 准备数据
        data = [
            {"姓名": "张三", "支行": "北京支行"},
            {"姓名": "李四", "支行": "上海支行"},
        ]

        # 字段映射（现在直接使用 field_mappings 处理支行）
        field_mappings = {
            "姓名": {"source_column": "姓名"},
            "部门": {"source_column": "支行"},  # 将支行映射到部门列
        }

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )

        # 验证结果
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(2, 2).value == "北京支行"
        assert ws_result.cell(3, 1).value == "李四"
        assert ws_result.cell(3, 2).value == "上海支行"

        wb_result.close()

    def test_month_type_mapping_month_number(self, tmp_path):
        """测试月类型映射功能 - 月份数字"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "月份"]])

        # 准备数据
        data = [{"姓名": "张三"}, {"姓名": "李四"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 月类型映射
        month_type_mapping = {"enabled": True, "target_column": "B"}

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件（使用月份参数）
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            month_type_mapping=month_type_mapping,
            month_param="1",
        )

        # 验证月类型映射
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(2, 2).value == "01月收入"
        assert ws_result.cell(3, 1).value == "李四"
        assert ws_result.cell(3, 2).value == "01月收入"

        wb_result.close()

    def test_month_type_mapping_bonus(self, tmp_path):
        """测试月类型映射功能 - 年终奖"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "类型"]])

        # 准备数据
        data = [{"姓名": "张三"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 月类型映射
        month_type_mapping = {
            "enabled": True,
            "target_column": "B",
            "bonus_value": "年终奖",
        }

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件（使用年终奖参数）
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            month_type_mapping=month_type_mapping,
            month_param="年终奖",
        )

        # 验证月类型映射
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(2, 2).value == "年终奖"

        wb_result.close()

    def test_column_index_mapping(self, tmp_path):
        """测试使用列索引的映射模式"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "年龄"]])

        # 准备数据
        data = [{"name": "张三", "age": 25}, {"name": "李四", "age": 30}]

        # 字段映射（使用列索引）
        field_mappings = {
            "name": {"source_column": "name", "target_column": "A"},
            "age": {"source_column": "age", "target_column": "B"},
        }

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_index",
        )

        # 验证数据
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(2, 2).value == 25
        assert ws_result.cell(3, 1).value == "李四"
        assert ws_result.cell(3, 2).value == 30

        wb_result.close()

    def test_old_format_field_mappings_xlsx(self, tmp_path):
        """测试旧格式field_mappings在xlsx中可用"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名", "年龄"]])

        data = [{"name": "张三", "age": 25}, {"name": "李四", "age": 30}]

        # 旧格式：输入列名 -> 模板列名
        field_mappings = {"name": "姓名", "age": "年龄"}

        output_path = tmp_path / "output.xlsx"
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )

        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(2, 1).value == "张三"
        assert ws_result.cell(2, 2).value == 25
        assert ws_result.cell(3, 1).value == "李四"
        assert ws_result.cell(3, 2).value == 30
        wb_result.close()

    def test_clear_existing_data(self, tmp_path):
        """测试清除现有数据"""
        template_path = write_xlsx_rows(
            tmp_path / "template.xlsx",
            [["姓名"], ["旧数据1"], ["旧数据2"], ["旧数据3"]],
        )

        # 准备新数据
        data = [{"姓名": "新数据1"}, {"姓名": "新数据2"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )

        # 验证旧数据被清除
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(1, 1).value == "姓名"
        assert ws_result.cell(2, 1).value == "新数据1"
        assert ws_result.cell(3, 1).value == "新数据2"
        assert ws_result.cell(4, 1).value is None  # 旧数据被清除

        wb_result.close()

    def test_config_validation_invalid_start_row(self, tmp_path):
        """测试配置验证 - 无效的start_row"""
        template_path = write_xlsx_rows(tmp_path / "template.xlsx", [["姓名"]])

        # 准备数据
        data = [{"姓名": "张三"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件（start_row <= header_row，应该抛出异常）
        writer = ExcelWriter()
        with pytest.raises(Exception):  # ConfigError
            writer.write_excel(
                template_path=str(template_path),
                data=data,
                field_mappings=field_mappings,
                output_path=str(output_path),
                header_row=2,
                start_row=2,  # 无效：start_row 必须大于 header_row
                mapping_mode="column_name",
            )

    def test_unsupported_format(self, tmp_path):
        """测试不支持的文件格式"""
        # 创建不支持的格式文件
        template_path = tmp_path / "template.txt"
        template_path.write_text("test data", encoding="utf-8")

        # 准备数据
        data = [{"姓名": "张三"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 输出路径
        output_path = tmp_path / "output.txt"

        # 写入文件（应该抛出异常）
        writer = ExcelWriter()
        with pytest.raises(ExcelError):
            writer.write_excel(
                template_path=str(template_path),
                data=data,
                field_mappings=field_mappings,
                output_path=str(output_path),
                header_row=1,
                start_row=2,
                mapping_mode="column_name",
            )

    def test_template_not_found(self, tmp_path):
        """测试模板文件未找到"""
        # 准备数据
        data = [{"姓名": "张三"}]

        # 字段映射
        field_mappings = {"姓名": {"source_column": "姓名"}}

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件（模板文件不存在）
        writer = ExcelWriter()
        with pytest.raises(FileNotFoundError):
            writer.write_excel(
                template_path="nonexistent.xlsx",
                data=data,
                field_mappings=field_mappings,
                output_path=str(output_path),
                header_row=1,
                start_row=2,
                mapping_mode="column_name",
            )

    def test_preserve_header_format_xlsx(self, tmp_path):
        """测试保留.xlsx文件的表头格式"""
        # 创建模板文件并设置格式
        template_path = tmp_path / "template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active

        # 设置表头样式（加粗）
        from openpyxl.styles import Font

        header_font = Font(bold=True)

        ws.cell(1, 1, "姓名").font = header_font
        ws.cell(1, 2, "年龄").font = header_font
        ws.cell(2, 1, "旧数据")
        wb.save(template_path)

        # 准备数据
        data = [{"姓名": "张三", "年龄": 25}]

        # 字段映射
        field_mappings = {
            "姓名": {"source_column": "姓名"},
            "年龄": {"source_column": "年龄"},
        }

        # 输出路径
        output_path = tmp_path / "output.xlsx"

        # 写入文件
        writer = ExcelWriter()
        writer.write_excel(
            template_path=str(template_path),
            data=data,
            field_mappings=field_mappings,
            output_path=str(output_path),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )

        # 验证表头格式被保留
        wb_result = openpyxl.load_workbook(output_path)
        ws_result = wb_result.active
        assert ws_result.cell(1, 1).font.bold  # 表头格式被保留
        assert ws_result.cell(1, 2).font.bold  # 表头格式被保留

        wb_result.close()
