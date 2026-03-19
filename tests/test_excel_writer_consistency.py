"""ExcelWriter 跨格式一致性测试。"""

from __future__ import annotations

import openpyxl
import pytest

from bank_template_processing.excel_writer import ExcelWriter
from tests.spreadsheet_factories import write_xls_rows, write_xlsx_rows


@pytest.mark.parametrize("ext", [".xlsx", ".xls"])
def test_writer_consistent_mapping_features(tmp_path, ext):
    template_path = tmp_path / f"template{ext}"
    output_path = tmp_path / f"output{ext}"

    if ext == ".xlsx":
        write_xlsx_rows(template_path, [["序号", "姓名", "金额", "用途", "备注"]])
    else:
        write_xls_rows(template_path, [["序号", "姓名", "金额", "用途", "备注"]])

    ExcelWriter().write_excel(
        template_path=str(template_path),
        data=[{"姓名": "张三", "金额": "12.5"}],
        field_mappings={
            "姓名": {"source_column": "姓名", "target_column": "姓名"},
            "金额": {"source_column": "金额", "target_column": "金额", "transform": "amount_decimal"},
        },
        output_path=str(output_path),
        header_row=1,
        start_row=2,
        mapping_mode="column_name",
        fixed_values={"备注": "固定"},
        auto_number={"enabled": True, "column_name": "序号", "start_from": 1},
        month_type_mapping={"enabled": True, "target_column": "用途", "month_format": "{month}月收入"},
        month_param="01",
        clear_rows={"start_row": 2, "end_row": 3},
    )

    if ext == ".xlsx":
        result = openpyxl.load_workbook(output_path)
        sheet = result.active
        assert sheet is not None
        assert sheet.cell(2, 1).value == 1
        assert sheet.cell(2, 2).value == "张三"
        assert sheet.cell(2, 3).value == 12.5
        assert sheet.cell(2, 4).value == "01月收入"
        assert sheet.cell(2, 5).value == "固定"
        result.close()
    else:
        import xlrd

        result = xlrd.open_workbook(str(output_path))
        sheet = result.sheet_by_index(0)
        assert sheet.cell_value(1, 0) == 1
        assert sheet.cell_value(1, 1) == "张三"
        assert sheet.cell_value(1, 2) == "12.5"
        assert sheet.cell_value(1, 3) == "01月收入"
        assert sheet.cell_value(1, 4) == "固定"
