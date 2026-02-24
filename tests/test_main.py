"""main.py 模块的测试"""

import json
import sys
from decimal import Decimal
from unittest.mock import MagicMock, patch

import pytest


class TestMonthValidation:
    """测试月份参数校验"""

    def test_validate_month_numeric_1_to_12(self):
        """测试数字格式月份：1-12"""
        from bank_template_processing.main import validate_month

        for month in range(1, 13):
            assert validate_month(str(month)) == str(month)

    def test_validate_month_numeric_with_leading(self):
        """测试数字格式月份：01-09"""
        from bank_template_processing.main import validate_month

        for month in range(1, 10):
            month_str = f"0{month}"
            assert validate_month(month_str) == month_str

    def test_validate_month_keyword_bonus(self):
        """测试关键字格式：年终奖"""
        from bank_template_processing.main import validate_month

        assert validate_month("年终奖") == "年终奖"

    def test_validate_month_keyword_compensation(self):
        """测试关键字格式：补偿金"""
        from bank_template_processing.main import validate_month

        assert validate_month("补偿金") == "补偿金"

    def test_validate_month_invalid_numeric(self):
        """测试无效的数字月份"""
        from bank_template_processing.main import validate_month

        with pytest.raises(ValueError, match="月份参数必须是1-12的数字、01-09格式、'年终奖'或'补偿金'"):
            validate_month("0")

        with pytest.raises(ValueError, match="月份参数必须是1-12的数字、01-09格式、'年终奖'或'补偿金'"):
            validate_month("13")

        with pytest.raises(ValueError, match="月份参数必须是1-12的数字、01-09格式、'年终奖'或'补偿金'"):
            validate_month("00")

    def test_validate_month_invalid_string(self):
        """测试无效的字符串月份"""
        from bank_template_processing.main import validate_month

        with pytest.raises(ValueError, match="月份参数必须是1-12的数字、01-09格式、'年终奖'或'补偿金'"):
            validate_month("1月")

        with pytest.raises(ValueError, match="月份参数必须是1-12的数字、01-09格式、'年终奖'或'补偿金'"):
            validate_month("January")


class TestArgumentParser:
    """测试命令行参数解析"""

    def test_parse_args_with_defaults(self):
        """测试使用默认值的参数解析"""
        from bank_template_processing.main import parse_args

        args = parse_args(["input.xlsx", "unit1", "01"])

        assert args.excel_path == "input.xlsx"
        assert args.unit_name == "unit1"
        assert args.month == "01"
        assert args.output_dir == "output/"
        assert args.config == "config.json"
        assert (
            args.output_filename_template
            == "{unit_name}_{template_name}_{count}人_金额{amount:.2f}元{ext}"
        )

    def test_parse_args_with_custom_values(self):
        """测试使用自定义值的参数解析"""
        from bank_template_processing.main import parse_args

        args = parse_args(
            [
                "input.xlsx",
                "unit1",
                "年终奖",
                "--output-dir",
                "custom_output/",
                "--config",
                "custom_config.json",
                "--output-filename-template",
                "{unit_name}_{month}_{timestamp}",
            ]
        )

        assert args.excel_path == "input.xlsx"
        assert args.unit_name == "unit1"
        assert args.month == "年终奖"
        assert args.output_dir == "custom_output/"
        assert args.config == "custom_config.json"
        assert args.output_filename_template == "{unit_name}_{month}_{timestamp}"


class TestGenerateOutputFilename:
    """测试输出文件名生成"""

    def test_generate_output_filename_without_template_name(self):
        """测试生成输出文件名（template_name为None时提取模板名）"""
        from bank_template_processing.main import generate_output_filename

        filename = generate_output_filename("unit1", "01", None, "template.xlsx", 10, 1000.50)

        assert filename == "unit1_template_10人_金额1000.50元.xlsx"

    def test_generate_output_filename_with_template_name(self):
        """测试生成输出文件名（显式提供template_name）"""
        from bank_template_processing.main import generate_output_filename

        filename = generate_output_filename("unit1", "年终奖", "工商银行", "template.xlsx", 5, 500.00)

        assert filename == "unit1_工商银行_5人_金额500.00元.xlsx"

    def test_generate_output_filename_csv_extension(self):
        """测试CSV模板生成CSV扩展名"""
        from bank_template_processing.main import generate_output_filename

        filename = generate_output_filename("unit1", "01", None, "template.csv", 10, 1000.00)

        assert filename == "unit1_template_10人_金额1000.00元.csv"

    def test_generate_output_filename_xls_extension(self):
        """测试XLS模板生成XLS扩展名"""
        from bank_template_processing.main import generate_output_filename

        filename = generate_output_filename("unit1", "01", None, "template.xls", 10, 1000.00)

        assert filename == "unit1_template_10人_金额1000.00元.xls"


class TestZeroSalaryFilterFunctions:
    """测试零工资筛选相关内部函数"""

    @pytest.mark.parametrize(
        ("value", "expected"),
        [
            (0, True),
            (0.0, True),
            (-0.0, True),
            (Decimal("0"), True),
            ("0", True),
            ("0.00", True),
            ("-0", True),
            ("0,000", True),
            ("0，000", True),
            (" 0 ", True),
            (1, False),
            ("1.00", False),
            (None, False),
            ("", False),
            ("N/A", False),
            (False, False),
            (True, False),
        ],
    )
    def test_is_zero_salary_value(self, value, expected):
        """测试零值判定"""
        from bank_template_processing.main import _is_zero_salary_value

        assert _is_zero_salary_value(value) is expected

    def test_filter_zero_salary_rows_missing_column_raises(self):
        """测试缺少实发工资列时抛异常"""
        from bank_template_processing.main import _filter_zero_salary_rows
        from bank_template_processing.validator import ValidationError

        with pytest.raises(ValidationError, match="缺少'实发工资'列"):
            _filter_zero_salary_rows([{"姓名": "张三"}, {"姓名": "李四"}])

    def test_filter_zero_salary_rows_filters_only_zero(self):
        """测试仅过滤零工资行"""
        from bank_template_processing.main import _filter_zero_salary_rows

        data = [
            {"姓名": "张三", "实发工资": 0},
            {"姓名": "李四", "实发工资": "0.00"},
            {"姓名": "王五", "实发工资": "1000.50"},
            {"姓名": "赵六", "实发工资": "N/A"},
            {"姓名": "孙七", "实发工资": None},
            {"姓名": "周八"},
        ]

        result = _filter_zero_salary_rows(data)

        assert [row["姓名"] for row in result] == ["王五", "赵六", "孙七", "周八"]


class TestMainWorkflow:
    """测试主工作流程"""

    def create_test_config(self, tmp_path, template_path=None):
        """创建测试配置文件"""
        config = {
            "version": "1.0",
            "organization_units": {
                "unit1": {
                    "template_path": template_path or str(tmp_path / "template.xlsx"),
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {},
                    "transformations": {},
                    "validation_rules": {},
                    "fixed_values": {},
                    "auto_number": {"enabled": False},
                    "bank_branch_mapping": {"enabled": False},
                    "month_type_mapping": {"enabled": False},
                }
            },
        }

        config_path = tmp_path / "config.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        return config_path

    def create_test_excel(self, tmp_path, filename="input.xlsx"):
        """创建测试Excel文件"""
        import openpyxl

        excel_path = tmp_path / filename
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None

        ws.append(["姓名", "卡号", "金额"])
        ws.append(["张三", "6222021234567890128", "1000.00"])

        wb.save(excel_path)
        return excel_path

    def create_template_excel(self, tmp_path, filename="template.xlsx"):
        """创建模板Excel文件"""
        import openpyxl

        template_path = tmp_path / filename
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None

        ws.append(["姓名", "卡号", "金额"])
        ws.append(["示例姓名", "示例卡号", "100.00"])

        wb.save(template_path)
        return template_path

    @patch("bank_template_processing.main.ExcelReader")
    @patch("bank_template_processing.main.ExcelWriter")
    def test_main_without_template_selection(self, mock_writer_class, mock_reader_class, tmp_path, caplog):
        """测试主流程：不启用模板选择"""
        from bank_template_processing.main import main

        input_excel = self.create_test_excel(tmp_path, "input.xlsx")
        template_excel = self.create_template_excel(tmp_path, "template.xlsx")
        config_path = self.create_test_config(tmp_path, str(template_excel))

        output_dir = tmp_path / "output"
        output_dir.mkdir()

        mock_reader_instance = MagicMock()
        mock_reader_instance.read_excel.return_value = [
            {
                "姓名": "张三",
                "卡号": "6222021234567890128",
                "金额": "1000.00",
                "实发工资": "1000.00",
            }
        ]
        mock_reader_class.return_value = mock_reader_instance

        mock_writer_instance = MagicMock()
        mock_writer_class.return_value = mock_writer_instance

        with patch.object(
            sys,
            "argv",
            [
                "main.py",
                str(input_excel),
                "unit1",
                "01",
                "--output-dir",
                str(output_dir),
                "--config",
                str(config_path),
            ],
        ):
            main()

        mock_reader_class.assert_called_once()
        mock_reader_instance.read_excel.assert_called_once()
        mock_writer_class.assert_called_once()

    @patch("bank_template_processing.main.ExcelReader")
    @patch("bank_template_processing.main.TemplateSelector")
    @patch("bank_template_processing.main.ExcelWriter")
    def test_main_with_template_selection(
        self,
        mock_writer_class,
        mock_selector_class,
        mock_reader_class,
        tmp_path,
        caplog,
    ):
        """测试主流程：启用模板选择"""
        from bank_template_processing.main import main

        input_excel = self.create_test_excel(tmp_path, "input.xlsx")
        template_excel = self.create_template_excel(tmp_path, "template.xlsx")
        special_template = self.create_template_excel(tmp_path, "special_template.xlsx")

        config = {
            "version": "1.0",
            "organization_units": {
                "unit1": {
                    "template_path": str(template_excel),
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {},
                    "transformations": {},
                    "validation_rules": {},
                    "fixed_values": {},
                    "auto_number": {"enabled": False},
                    "bank_branch_mapping": {"enabled": False},
                    "month_type_mapping": {"enabled": False},
                }
            },
            "template_selection_rules": {
                "enabled": True,
                "bank_column": "开户银行",
                "default_bank": "工商银行",
                "default_template": str(template_excel),
                "special_template": str(special_template),
            },
        }

        config_path = tmp_path / "config.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        output_dir = tmp_path / "output"
        output_dir.mkdir()

        mock_reader_instance = MagicMock()
        mock_reader_instance.read_excel.return_value = [
            {
                "姓名": "张三",
                "卡号": "6222021234567890128",
                "金额": "1000.00",
                "实发工资": "0",
                "开户银行": "工商银行",
            },
            {
                "姓名": "李四",
                "卡号": "6222021234567890128",
                "金额": "2000.00",
                "实发工资": "2000.00",
                "开户银行": "工商银行",
            }
        ]
        mock_reader_class.return_value = mock_reader_instance

        mock_selector_instance = MagicMock()
        mock_selector_instance.group_data.return_value = {
            "default": {
                "data": [
                    {
                        "姓名": "李四",
                        "卡号": "6222021234567890128",
                        "金额": "2000.00",
                        "实发工资": "2000.00",
                        "开户银行": "工商银行",
                    }
                ],
                "template": str(template_excel),
                "group_name": "template",
            },
            "special": {
                "data": [],
                "template": str(special_template),
                "group_name": "special_template",
            },
        }
        mock_selector_class.return_value = mock_selector_instance

        mock_writer_instance = MagicMock()
        mock_writer_class.return_value = mock_writer_instance

        with patch.object(
            sys,
            "argv",
            [
                "main.py",
                str(input_excel),
                "unit1",
                "01",
                "--output-dir",
                str(output_dir),
                "--config",
                str(config_path),
            ],
        ):
            main()

        mock_reader_class.assert_called_once()
        mock_selector_class.assert_called_once()
        mock_selector_instance.group_data.assert_called_once()
        grouped_data = mock_selector_instance.group_data.call_args.args[0]
        assert len(grouped_data) == 1
        assert grouped_data[0]["姓名"] == "李四"
        mock_writer_class.assert_called_once()

    def test_main_file_not_found(self, tmp_path, caplog):
        """测试主流程：输入文件不存在"""
        from bank_template_processing.main import main

        config = {
            "version": "1.0",
            "organization_units": {
                "unit1": {
                    "template_path": "template.xlsx",
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {},
                    "transformations": {},
                    "validation_rules": {},
                    "fixed_values": {},
                    "auto_number": {"enabled": False},
                    "bank_branch_mapping": {"enabled": False},
                    "month_type_mapping": {"enabled": False},
                }
            },
        }

        config_path = tmp_path / "config.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        with (
            patch("sys.exit") as mock_exit,
            patch.object(
                sys,
                "argv",
                [
                    "main.py",
                    "nonexistent.xlsx",
                    "unit1",
                    "01",
                    "--config",
                    str(config_path),
                ],
            ),
        ):
            main()

        mock_exit.assert_called_once_with(1)
        assert any("错误" in record.message for record in caplog.records)


class TestMainZeroSalaryFiltering:
    """测试主流程中的零工资筛选行为"""

    def _write_basic_config(self, tmp_path, template_path: str) -> str:
        config = {
            "version": "1.0",
            "organization_units": {
                "unit1": {
                    "template_path": template_path,
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {},
                    "transformations": {},
                    "validation_rules": {},
                }
            },
        }

        config_path = tmp_path / "config_basic.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return str(config_path)

    @patch("bank_template_processing.main.ExcelReader")
    @patch("bank_template_processing.main.ExcelWriter")
    def test_filter_before_validation_transform_and_write(
        self,
        mock_writer_class,
        mock_reader_class,
        tmp_path,
        caplog,
    ):
        """测试零工资行在验证/转换/写入前被过滤"""
        from bank_template_processing.main import main

        config = {
            "version": "1.0",
            "organization_units": {
                "unit1": {
                    "template_path": str(tmp_path / "template.xlsx"),
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {
                        "金额": {
                            "source_column": "实发工资",
                            "transform": "amount_decimal",
                        }
                    },
                    "transformations": {
                        "amount_decimal": {
                            "decimal_places": 2,
                        }
                    },
                    "validation_rules": {
                        "required_fields": ["姓名"],
                    },
                }
            },
        }
        config_path = tmp_path / "config_transform.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)

        mock_reader_instance = MagicMock()
        mock_reader_instance.read_excel.return_value = [
            {"姓名": "张三", "实发工资": "0"},
            {"姓名": "李四", "实发工资": "1000.456"},
        ]
        mock_reader_class.return_value = mock_reader_instance

        mock_writer_instance = MagicMock()
        mock_writer_class.return_value = mock_writer_instance

        caplog.set_level("INFO")
        with (
            patch("bank_template_processing.main.Validator.validate_required") as mock_validate_required,
            patch(
                "bank_template_processing.main.Transformer.transform_amount",
                return_value=1000.46,
            ) as mock_transform_amount,
            patch.object(
                sys,
                "argv",
                [
                    "main.py",
                    str(tmp_path / "input.xlsx"),
                    "unit1",
                    "01",
                    "--output-dir",
                    str(tmp_path / "output"),
                    "--config",
                    str(config_path),
                ],
            ),
        ):
            main()

        assert mock_validate_required.call_count == 1
        assert mock_transform_amount.call_count == 1
        assert mock_transform_amount.call_args.args[0] == "1000.456"

        written_data = mock_writer_instance.write_excel.call_args.kwargs["data"]
        assert len(written_data) == 1
        assert written_data[0]["姓名"] == "李四"
        assert written_data[0]["实发工资"] == 1000.46
        assert any("原始 2 行，过滤 1 行，保留 1 行" in record.message for record in caplog.records)

    @patch("bank_template_processing.main.ExcelReader")
    def test_main_missing_salary_column_exits(self, mock_reader_class, tmp_path):
        """测试主流程缺少实发工资列时退出"""
        from bank_template_processing.main import main

        config_path = self._write_basic_config(tmp_path, str(tmp_path / "template.xlsx"))
        mock_reader_instance = MagicMock()
        mock_reader_instance.read_excel.return_value = [{"姓名": "张三"}]
        mock_reader_class.return_value = mock_reader_instance

        with (
            patch("sys.exit") as mock_exit,
            patch.object(
                sys,
                "argv",
                [
                    "main.py",
                    str(tmp_path / "input.xlsx"),
                    "unit1",
                    "01",
                    "--config",
                    config_path,
                ],
            ),
        ):
            main()

        mock_exit.assert_called_once_with(1)


class TestMainZeroSalaryFilteringByInputFormat:
    """测试零工资筛选对三种输入格式均生效"""

    def _create_input_file(self, tmp_path, suffix: str) -> str:
        file_path = tmp_path / f"input{suffix}"
        headers = ["姓名", "实发工资", "开户银行"]
        rows = [
            ["张三", "0", "工商银行"],
            ["李四", "1000.50", "工商银行"],
        ]

        if suffix == ".xlsx":
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            assert ws is not None
            ws.append(headers)
            for row in rows:
                ws.append(row)
            wb.save(file_path)
        elif suffix == ".csv":
            lines = [",".join(headers)] + [",".join(row) for row in rows]
            file_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        elif suffix == ".xls":
            import xlwt

            wb = xlwt.Workbook(encoding="utf-8")
            ws = wb.add_sheet("Sheet1")
            for col_idx, value in enumerate(headers):
                ws.write(0, col_idx, value)
            for row_idx, row in enumerate(rows, start=1):
                for col_idx, value in enumerate(row):
                    ws.write(row_idx, col_idx, value)
            wb.save(str(file_path))
        else:
            raise ValueError(f"不支持的测试后缀: {suffix}")

        return str(file_path)

    def _create_config(self, tmp_path, template_path: str) -> str:
        config = {
            "version": "1.0",
            "organization_units": {
                "unit1": {
                    "template_path": template_path,
                    "header_row": 1,
                    "start_row": 2,
                    "field_mappings": {},
                    "transformations": {},
                    "validation_rules": {},
                }
            },
        }
        config_path = tmp_path / "config_formats.json"
        with open(config_path, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return str(config_path)

    @pytest.mark.parametrize("suffix", [".xlsx", ".csv", ".xls"])
    @patch("bank_template_processing.main.ExcelWriter")
    def test_main_filters_zero_salary_for_all_input_formats(self, mock_writer_class, tmp_path, suffix):
        """测试 .xlsx/.csv/.xls 输入均会过滤零工资行"""
        from bank_template_processing.main import _is_zero_salary_value, main

        input_path = self._create_input_file(tmp_path, suffix)
        config_path = self._create_config(tmp_path, str(tmp_path / "template.xlsx"))
        output_dir = tmp_path / "output"
        output_dir.mkdir()

        mock_writer_instance = MagicMock()
        mock_writer_class.return_value = mock_writer_instance

        with patch.object(
            sys,
            "argv",
            [
                "main.py",
                input_path,
                "unit1",
                "01",
                "--output-dir",
                str(output_dir),
                "--config",
                config_path,
            ],
        ):
            main()

        written_data = mock_writer_instance.write_excel.call_args.kwargs["data"]
        assert len(written_data) == 1
        assert written_data[0]["姓名"] == "李四"
        assert not _is_zero_salary_value(written_data[0].get("实发工资"))


class TestApplyTransformations:
    """测试数据转换"""

    def test_apply_transformations_empty(self):
        """测试空转换配置"""
        from bank_template_processing.main import apply_transformations

        data = [{"姓名": "张三"}]
        result = apply_transformations(data, {}, {})

        assert result == data

    def test_apply_transformations_date(self):
        """测试日期转换"""
        from bank_template_processing.main import apply_transformations

        data = [{"日期": "15/01/2024"}]
        transformations = {"date_format": {"output_format": "YYYY-MM-DD"}}
        field_mappings = {"日期": {"source_column": "日期", "transform": "date_format"}}
        result = apply_transformations(data, transformations, field_mappings)

        assert result[0]["日期"] == "2024-01-15"

    def test_apply_transformations_amount(self):
        """测试金额转换"""
        from bank_template_processing.main import apply_transformations

        data = [{"金额": "1000.456"}]
        transformations = {"amount_decimal": {"decimal_places": 2}}
        field_mappings = {"金额": {"source_column": "金额", "transform": "amount_decimal"}}
        result = apply_transformations(data, transformations, field_mappings)

        assert result[0]["金额"] == 1000.46

    def test_apply_transformations_amount_zero(self):
        """测试金额为0时仍进行转换"""
        from bank_template_processing.main import apply_transformations

        data = [{"金额": 0}]
        transformations = {"amount_decimal": {"decimal_places": 2}}
        field_mappings = {"金额": {"source_column": "金额", "transform": "amount_decimal"}}
        result = apply_transformations(data, transformations, field_mappings)

        assert result[0]["金额"] == 0.0

    def test_apply_transformations_card_number(self):
        """测试卡号转换"""
        from bank_template_processing.main import apply_transformations

        data = [{"卡号": "6222-0212-3456-7890-128"}]
        transformations = {"card_number": {"remove_formatting": True, "luhn_validation": True}}
        field_mappings = {"卡号": {"source_column": "卡号", "transform": "card_number"}}
        result = apply_transformations(data, transformations, field_mappings)
        assert result[0]["卡号"] == "6222021234567890128"

    def test_apply_transformations_mixed(self):
        """测试混合转换"""
        from bank_template_processing.main import apply_transformations

        data = [{"日期": "15/01/2024", "金额": "1000.456"}]
        transformations = {
            "date_format": {"output_format": "YYYY-MM-DD"},
            "amount_decimal": {"decimal_places": 2},
        }
        field_mappings = {
            "日期": {"source_column": "日期", "transform": "date_format"},
            "金额": {"source_column": "金额", "transform": "amount_decimal"},
        }
        result = apply_transformations(data, transformations, field_mappings)

        assert result[0]["日期"] == "2024-01-15"
        assert result[0]["金额"] == 1000.46
