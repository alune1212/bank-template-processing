"""
集成测试 - 端到端工作流验证
"""

import json
import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook

# 添加项目根目录到Python路径
sys.path.insert(0, str(Path(__file__).parent.parent))

from config_loader import ConfigError, load_config, validate_config
from excel_reader import ExcelReader, ExcelError
from excel_writer import ExcelWriter
from template_selector import TemplateSelector, ValidationError
from transformer import Transformer, TransformError
from validator import Validator


class TestEndToEndWorkflow:
    """测试端到端完整工作流"""

    def test_complete_workflow(self, tmp_path):
        """完整工作流: 读取配置 → 读取输入 → 转换 → 写入输出"""

        input_path = "tests/fixtures/integration_input.xlsx"
        template_path = "tests/fixtures/integration_template.xlsx"
        config_path = "tests/fixtures/integration_config_compat.json"

        output_path = tmp_path / "output.xlsx"

        config = load_config(config_path)
        validate_config(config)

        unit_config = config["organization_units"]["测试单位"]

        reader = ExcelReader()
        data = reader.read_excel(input_path)

        transformer = Transformer()
        for row in data:
            for field_name, transform_config in unit_config["transformations"].items():
                if field_name in row:
                    transform_type = transform_config["type"]
                    if transform_type == "date":
                        row[field_name] = transformer.transform_date(row[field_name])
                    elif transform_type == "amount":
                        decimal_places = transform_config.get("decimal_places", 2)
                        result = transformer.transform_amount(
                            row[field_name], decimal_places
                        )
                        row[field_name] = float(result)
                    elif transform_type == "card_number":
                        row[field_name] = transformer.transform_card_number(
                            row[field_name]
                        )

        writer = ExcelWriter()

        simple_field_mappings = {
            k: v["column_name"] for k, v in unit_config["field_mappings"].items()
        }

        writer.write_excel(
            template_path=template_path,
            data=data,
            field_mappings=simple_field_mappings,
            output_path=str(output_path),
            header_row=unit_config["header_row"],
            start_row=unit_config["start_row"],
            mapping_mode="column_name",
            fixed_values=None,
            auto_number=unit_config.get("auto_number"),
            bank_branch_mapping=unit_config.get("bank_branch_mapping"),
            month_type_mapping=unit_config.get("month_type_mapping"),
            month_param="01",
        )

        assert output_path.exists()

        wb = load_workbook(output_path)
        ws = wb.active

        assert ws.max_row == 9

        assert ws.cell(row=4, column=2).value == "张三"
        assert ws.cell(row=4, column=3).value == "6222021234567890128"
        assert ws.cell(row=4, column=4).value == "北京分行"
        assert ws.cell(row=4, column=5).value == 1000.50
        assert ws.cell(row=4, column=6).value == "2025-01-15"

        assert ws.cell(row=5, column=1).value == 2
        assert ws.cell(row=6, column=1).value == 3
        assert ws.cell(row=7, column=1).value == 4


class TestErrorHandling:
    """测试错误处理"""

    def test_missing_input_file(self):
        """测试缺失输入文件"""
        with pytest.raises(FileNotFoundError):
            reader = ExcelReader()
            reader.read_excel("nonexistent.xlsx")

    def test_missing_template_file(self, tmp_path):
        """测试缺失模板文件"""
        output_path = tmp_path / "output.xlsx"

        with pytest.raises(FileNotFoundError):
            writer = ExcelWriter()
            writer.write_excel(
                template_path="tests/fixtures/nonexistent.xlsx",
                data=[],
                field_mappings={"姓名": {"column_name": "姓名"}},
                output_path=str(output_path),
                header_row=3,
                start_row=4,
                mapping_mode="column_name",
            )

    def test_invalid_config(self, tmp_path):
        """测试无效配置"""
        invalid_config = {
            "version": "1.0",
            "organization_units": {
                "测试单位": {
                    "header_row": 3,
                    "start_row": 4,
                    "field_mappings": {"姓名": {"column_name": "姓名"}},
                    "transformations": {},
                }
            },
        }

        with pytest.raises(ConfigError, match="缺少必填字段.*template_path"):
            validate_config(invalid_config)

    def test_invalid_card_number(self):
        """测试无效卡号转换"""
        transformer = Transformer()

        with pytest.raises(TransformError):
            transformer.transform_card_number("1234567890123456")


class TestDataTransformation:
    """测试数据转换验证"""

    def test_date_transformation(self):
        """测试日期转换"""
        transformer = Transformer()

        result = transformer.transform_date("2025-01-15")
        assert result == "2025-01-15"

        result = transformer.transform_date("15/01/2025")
        assert result == "2025-01-15"

        result = transformer.transform_date("01/15/2025")
        assert result == "2025-01-15"

    def test_amount_transformation(self):
        """测试金额转换"""
        transformer = Transformer()

        result = transformer.transform_amount("1000.50", 2)
        assert result == 1000.50

        result = transformer.transform_amount(2000, 2)
        assert result == 2000.00

    def test_card_number_transformation(self):
        """测试卡号转换"""
        transformer = Transformer()

        result = transformer.transform_card_number("6222 0212 3456 7890 128")
        assert result == "6222021234567890128"

        result = transformer.transform_card_number("6222-0212-3456-7890-128")
        assert result == "6222021234567890128"


class TestUniqueFilenameGeneration:
    """测试唯一文件名生成"""

    def test_filename_with_template_name(self):
        """测试带模板名的文件名生成"""
        from main import generate_output_filename, generate_timestamp

        timestamp = generate_timestamp()
        filename = generate_output_filename("测试单位", "01", "农业银行", timestamp)

        assert "测试单位" in filename
        assert "农业银行" in filename
        assert "01" in filename
        assert timestamp in filename
        assert filename.endswith(".xlsx")

    def test_filename_without_template_name(self):
        """测试不带模板名的文件名生成"""
        from main import generate_output_filename, generate_timestamp

        timestamp = generate_timestamp()
        filename = generate_output_filename("测试单位", "01", None, timestamp)

        assert "测试单位" in filename
        assert "01" in filename
        assert timestamp in filename
        assert filename.endswith(".xlsx")
        assert "农业银行" not in filename

    def test_timestamp_uniqueness(self):
        """测试时间戳唯一性"""
        from main import generate_timestamp
        import time

        timestamp1 = generate_timestamp()
        time.sleep(1)
        timestamp2 = generate_timestamp()

        assert timestamp1 != timestamp2


class TestLuhnValidation:
    """测试Luhn校验功能"""

    def test_valid_card_numbers(self):
        """测试有效卡号"""
        transformer = Transformer()

        valid_cards = [
            "4111111111111111",
            "6222021234567890128",
            "4012888888881881",
        ]

        for card in valid_cards:
            result = transformer.transform_card_number(card)
            assert result == card

    def test_invalid_card_numbers(self):
        """测试无效卡号"""
        transformer = Transformer()

        invalid_cards = [
            "4111111111111112",
            "1234567890123456",
        ]

        for card in invalid_cards:
            with pytest.raises(TransformError):
                transformer.transform_card_number(card)


class TestDynamicTemplateSelection:
    """测试动态模板选择功能"""

    def test_mixed_banks_generates_two_files(self, tmp_path):
        """混合银行 → 生成两个文件"""
        input_path = "tests/fixtures/integration_input.xlsx"
        config_path = "tests/fixtures/integration_config.json"

        config = load_config(config_path)
        validate_config(config)

        unit_config = config["organization_units"]["测试单位"]
        template_rules = config.get("template_selector", {})

        reader = ExcelReader()
        data = reader.read_excel(input_path)

        selector = TemplateSelector(config)
        if selector.is_enabled():
            groups = selector.group_data(
                data, default_bank=template_rules.get("default_bank")
            )
        else:
            groups = {
                "default": {"data": data, "template": unit_config["template_path"]}
            }

        from main import generate_timestamp

        timestamp = generate_timestamp()
        output_files = []

        for group_key in ["default", "special"]:
            group = groups.get(group_key, {})
            group_data = group.get("data", [])
            if not group_data:
                continue

            template_path = group["template"]
            group_name = group.get("group_name", "")

            transformer = Transformer()
            for row in group_data:
                for field_name, transform_config in unit_config[
                    "transformations"
                ].items():
                    if field_name in row:
                        transform_type = transform_config["type"]
                        if transform_type == "date":
                            row[field_name] = transformer.transform_date(
                                row[field_name]
                            )
                        elif transform_type == "amount":
                            decimal_places = transform_config.get("decimal_places", 2)
                            row[field_name] = transformer.transform_amount(
                                row[field_name], decimal_places
                            )
                        elif transform_type == "card_number":
                            row[field_name] = transformer.transform_card_number(
                                row[field_name]
                            )

            from main import generate_output_filename

            filename = generate_output_filename("测试单位", "01", group_name, timestamp)
            output_path = tmp_path / filename

            simple_field_mappings = {
                k: v["column_name"] for k, v in unit_config["field_mappings"].items()
            }

            writer = ExcelWriter()
            writer.write_excel(
                template_path=template_path,
                data=group_data,
                field_mappings=simple_field_mappings,
                output_path=str(output_path),
                header_row=unit_config["header_row"],
                start_row=unit_config["start_row"],
                mapping_mode="column_name",
                fixed_values=None,
                auto_number=unit_config.get("auto_number"),
                bank_branch_mapping=unit_config.get("bank_branch_mapping"),
                month_type_mapping=unit_config.get("month_type_mapping"),
                month_param="01",
            )

            output_files.append(output_path)

        assert len(output_files) == 2

        default_file = [
            f
            for f in output_files
            if "integration_template" in f.name and "special" not in f.name
        ][0]
        special_file = [f for f in output_files if "special" in f.name][0]

        wb_default = load_workbook(default_file)
        ws_default = wb_default.active
        assert ws_default.max_row == 6

        wb_special = load_workbook(special_file)
        ws_special = wb_special.active
        assert ws_special.max_row == 6

    def test_all_default_bank_generates_one_file(self, tmp_path):
        """全默认银行 → 生成一个文件（默认模板）"""

        config = {
            "template_selection_rules": {
                "enabled": True,
                "default_bank": "农业银行",
                "default_template": "tests/fixtures/integration_template.xlsx",
                "special_template": "tests/fixtures/integration_special_template.xlsx",
            }
        }

        data = [
            {"开户银行": "农业银行", "姓名": "张三", "金额": 1000},
            {"开户银行": "农业银行", "姓名": "王五", "金额": 3000},
        ]

        selector = TemplateSelector(config)
        groups = selector.group_data(data, default_bank="农业银行")

        assert len(groups["default"]["data"]) == 2
        assert len(groups["special"]["data"]) == 0

    def test_all_special_bank_generates_one_file(self, tmp_path):
        """全特殊银行 → 生成一个文件（特殊模板）"""

        config = {
            "template_selection_rules": {
                "enabled": True,
                "default_bank": "农业银行",
                "default_template": "tests/fixtures/integration_template.xlsx",
                "special_template": "tests/fixtures/integration_special_template.xlsx",
            }
        }

        data = [
            {"开户银行": "工商银行", "姓名": "李四", "金额": 2000},
            {"开户银行": "建设银行", "姓名": "赵六", "金额": 4000},
        ]

        selector = TemplateSelector(config)
        groups = selector.group_data(data, default_bank="农业银行")

        assert len(groups["default"]["data"]) == 0
        assert len(groups["special"]["data"]) == 2

    def test_no_template_selection_rules_single_file(self):
        """无template_selection_rules → 单文件（向后兼容）"""

        config = {}

        selector = TemplateSelector(config)
        assert selector.is_enabled() is False

    def test_enabled_false_single_file(self):
        """enabled=false → 单文件（向后兼容）"""

        config = {
            "template_selection_rules": {
                "enabled": False,
                "default_bank": "农业银行",
            }
        }

        selector = TemplateSelector(config)
        assert selector.is_enabled() is False

    def test_missing_bank_column_raises_error(self):
        """开户银行列不存在 → ValidationError"""

        config = {
            "template_selection_rules": {
                "enabled": True,
                "default_bank": "农业银行",
                "default_template": "tests/fixtures/integration_template.xlsx",
                "special_template": "tests/fixtures/integration_special_template.xlsx",
                "bank_column": "开户银行",
            }
        }

        data = [
            {"姓名": "张三", "金额": 1000},
            {"姓名": "李四", "金额": 2000},
        ]

        selector = TemplateSelector(config)
        with pytest.raises(ValidationError, match="缺少'开户银行'列"):
            selector.group_data(data, default_bank="农业银行")

    def test_empty_bank_value_raises_error(self):
        """开户银行为空 → ValidationError"""

        config = {
            "template_selection_rules": {
                "enabled": True,
                "default_bank": "农业银行",
                "default_template": "tests/fixtures/integration_template.xlsx",
                "special_template": "tests/fixtures/integration_special_template.xlsx",
                "bank_column": "开户银行",
            }
        }

        data = [
            {"开户银行": "农业银行", "姓名": "张三", "金额": 1000},
            {"开户银行": "", "姓名": "李四", "金额": 2000},
        ]

        selector = TemplateSelector(config)
        with pytest.raises(ValidationError, match="第2行的'开户银行'字段为空"):
            selector.group_data(data, default_bank="农业银行")

    def test_empty_group_skips_output(self):
        """空组跳过 → 不生成对应文件"""

        config = {
            "template_selection_rules": {
                "enabled": True,
                "default_bank": "农业银行",
                "default_template": "tests/fixtures/integration_template.xlsx",
                "special_template": "tests/fixtures/integration_special_template.xlsx",
            }
        }

        data = [
            {"开户银行": "农业银行", "姓名": "张三", "金额": 1000},
            {"开户银行": "农业银行", "姓名": "王五", "金额": 3000},
        ]

        selector = TemplateSelector(config)
        groups = selector.group_data(data, default_bank="农业银行")

        assert len(groups["default"]["data"]) == 2
        assert len(groups["special"]["data"]) == 0
