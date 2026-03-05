"""ExcelWriter 失败路径与边界分支测试。"""

from __future__ import annotations

import csv
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

import bank_template_processing.excel_writer as excel_writer_module
from bank_template_processing.config_loader import ConfigError
from bank_template_processing.excel_writer import ExcelError, ExcelWriter


def _create_csv_template(path: Path) -> None:
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["姓名", "金额"])


def _create_xlsx_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.cell(1, 1, "姓名")
    ws.cell(1, 2, "金额")
    wb.save(path)


def test_write_excel_wraps_unexpected_exception(tmp_path, monkeypatch):
    template_path = tmp_path / "template.csv"
    _create_csv_template(template_path)

    writer = ExcelWriter()

    def boom(*_args, **_kwargs):
        raise RuntimeError("unexpected")

    monkeypatch.setattr(writer, "_write_csv", boom)

    with pytest.raises(ExcelError, match="写入文件失败"):
        writer.write_excel(
            template_path=str(template_path),
            data=[{"姓名": "张三"}],
            field_mappings={"姓名": {"source_column": "姓名"}},
            output_path=str(tmp_path / "out.csv"),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )


def test_write_csv_header_row_out_of_range_raises(tmp_path):
    template_path = tmp_path / "template.csv"
    template_path.write_text("姓名,金额\n", encoding="utf-8")

    with pytest.raises(ExcelError, match="模板文件行数不足"):
        ExcelWriter().write_excel(
            template_path=str(template_path),
            data=[{"姓名": "张三"}],
            field_mappings={"姓名": {"source_column": "姓名"}},
            output_path=str(tmp_path / "out.csv"),
            header_row=5,
            start_row=6,
            mapping_mode="column_name",
        )


def test_write_csv_clear_rows_invalid_range_raises(tmp_path):
    template_path = tmp_path / "template.csv"
    _create_csv_template(template_path)

    with pytest.raises(ExcelError, match="start_row 不能大于 end_row"):
        ExcelWriter().write_excel(
            template_path=str(template_path),
            data=[{"姓名": "张三"}],
            field_mappings={"姓名": {"source_column": "姓名"}},
            output_path=str(tmp_path / "out.csv"),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
            clear_rows={"start_row": 5, "end_row": 2},
        )


def test_write_xlsx_openpyxl_missing_raises(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xlsx"
    _create_xlsx_template(template_path)

    monkeypatch.setattr(excel_writer_module, "openpyxl", None)

    with pytest.raises(ExcelError, match="openpyxl未安装"):
        ExcelWriter().write_excel(
            template_path=str(template_path),
            data=[{"姓名": "张三"}],
            field_mappings={"姓名": {"source_column": "姓名"}},
            output_path=str(tmp_path / "out.xlsx"),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )


def test_write_xlsx_invalid_file_exception_raises(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xlsx"
    _create_xlsx_template(template_path)

    class FakeOpenpyxl:
        @staticmethod
        def load_workbook(*_args, **_kwargs):
            raise excel_writer_module.InvalidFileException("bad file")

    monkeypatch.setattr(excel_writer_module, "openpyxl", FakeOpenpyxl)

    with pytest.raises(ExcelError, match="无效的Excel文件"):
        ExcelWriter()._write_xlsx(
            template_path=str(template_path),
            data=[],
            field_mappings={},
            output_path=str(tmp_path / "out.xlsx"),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )


def test_write_xlsx_without_worksheet_raises(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xlsx"
    _create_xlsx_template(template_path)

    class FakeWorkbook:
        worksheets = []
        active = None

        def save(self, _output_path):
            raise AssertionError("should not save when no worksheet")

    monkeypatch.setattr(
        excel_writer_module,
        "openpyxl",
        SimpleNamespace(load_workbook=lambda *_args, **_kwargs: FakeWorkbook()),
    )

    with pytest.raises(ExcelError, match="模板文件没有工作表"):
        ExcelWriter()._write_xlsx(
            template_path=str(template_path),
            data=[],
            field_mappings={},
            output_path=str(tmp_path / "out.xlsx"),
            header_row=1,
            start_row=2,
            mapping_mode="column_name",
        )


def test_write_xls_missing_dependencies_raises(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xls"
    template_path.write_text("dummy", encoding="utf-8")

    monkeypatch.setattr(excel_writer_module, "xlwt", None)
    with pytest.raises(ExcelError, match="xlwt未安装"):
        ExcelWriter()._write_xls(str(template_path), [], {}, str(tmp_path / "out.xls"), 1, 2, "column_name")

    monkeypatch.setattr(excel_writer_module, "xlwt", object())
    monkeypatch.setattr(excel_writer_module, "xlrd", None)
    with pytest.raises(ExcelError, match="xlrd未安装"):
        ExcelWriter()._write_xls(str(template_path), [], {}, str(tmp_path / "out.xls"), 1, 2, "column_name")

    monkeypatch.setattr(excel_writer_module, "xlrd", object())
    monkeypatch.setattr(excel_writer_module, "xl_copy", None)
    with pytest.raises(ExcelError, match="xlutils未安装"):
        ExcelWriter()._write_xls(str(template_path), [], {}, str(tmp_path / "out.xls"), 1, 2, "column_name")


def test_write_xls_open_workbook_failure_raises(tmp_path, monkeypatch):
    template_path = tmp_path / "template.xls"
    template_path.write_text("dummy", encoding="utf-8")

    fake_xlrd = SimpleNamespace(
        open_workbook=lambda *_args, **_kwargs: (_ for _ in ()).throw(RuntimeError("open failed"))
    )

    monkeypatch.setattr(excel_writer_module, "xlwt", object())
    monkeypatch.setattr(excel_writer_module, "xlrd", fake_xlrd)
    monkeypatch.setattr(excel_writer_module, "xl_copy", lambda _wb: None)

    with pytest.raises(ExcelError, match="无法读取Excel文件"):
        ExcelWriter()._write_xls(str(template_path), [], {}, str(tmp_path / "out.xls"), 1, 2, "column_name")


def test_calculate_month_value_error_paths():
    writer = ExcelWriter()

    with pytest.raises(ConfigError, match="month_format 缺少变量"):
        writer._calculate_month_value("1", {"month_format": "{missing}"})

    with pytest.raises(ConfigError, match="month_format 格式错误"):
        writer._calculate_month_value("1", {"month_format": "{month"})

    assert writer._calculate_month_value("未知", {"month_format": "{month}月收入"}) is None
    assert writer._calculate_month_value(None, {"month_format": "{month}月收入"}) is None


def test_resolve_column_index_error_paths():
    writer = ExcelWriter()

    with pytest.raises(ValueError, match="列索引必须 >= 1"):
        writer._resolve_column_index(0)

    with pytest.raises(ValueError, match="列索引必须 >= 1"):
        writer._resolve_column_index("0")

    with pytest.raises(ValueError, match="超出最大列数"):
        writer._resolve_column_index("B", max_columns=1, strict_bounds=True)

    with pytest.raises(ValueError, match="无法解析列标识"):
        writer._resolve_column_index("中文列")


def test_resolve_column_index_by_mode_fallback_to_headers(caplog):
    writer = ExcelWriter()
    caplog.set_level("WARNING")

    col_idx = writer._resolve_column_index_by_mode(
        "姓名",
        headers={"姓名": 1},
        max_columns=2,
        mapping_mode="column_index",
    )

    assert col_idx == 1
    assert "回退按表头解析" in caplog.text


def test_resolve_column_index_by_mode_without_headers_raises():
    writer = ExcelWriter()
    with pytest.raises(ValueError, match="无法解析列标识"):
        writer._resolve_column_index_by_mode("中文列", headers=None, max_columns=3, mapping_mode="column_index")


def test_write_excel_warns_on_deprecated_bank_branch_mapping(tmp_path, caplog):
    template_path = tmp_path / "template.csv"
    _create_csv_template(template_path)

    caplog.set_level("WARNING")
    ExcelWriter().write_excel(
        template_path=str(template_path),
        data=[],
        field_mappings={},
        output_path=str(tmp_path / "out.csv"),
        header_row=1,
        start_row=2,
        mapping_mode="column_name",
        bank_branch_mapping={"enabled": True},
    )
    assert "bank_branch_mapping" in caplog.text


def test_write_xlsx_header_row_zero_with_invalid_clear_range_raises(tmp_path):
    template_path = tmp_path / "template.xlsx"
    _create_xlsx_template(template_path)

    with pytest.raises(ConfigError, match="clear_rows.start_row 不能大于 end_row"):
        ExcelWriter()._write_xlsx(
            template_path=str(template_path),
            data=[],
            field_mappings={},
            output_path=str(tmp_path / "out.xlsx"),
            header_row=0,
            start_row=2,
            mapping_mode="column_name",
            clear_rows={"start_row": 3, "end_row": 2},
        )


def test_write_xls_header_row_zero_and_clear_rows_success(tmp_path):
    import xlwt

    template_path = tmp_path / "template.xls"
    workbook = xlwt.Workbook(encoding="utf-8")
    sheet = workbook.add_sheet("Sheet1")
    sheet.write(0, 0, "说明")
    sheet.write(1, 0, "姓名")
    sheet.write(2, 0, "旧数据")
    workbook.save(template_path)

    ExcelWriter()._write_xls(
        template_path=str(template_path),
        data=[{"姓名": "张三"}],
        field_mappings={"姓名": {"source_column": "姓名", "target_column": "A"}},
        output_path=str(tmp_path / "out.xls"),
        header_row=0,
        start_row=2,
        mapping_mode="column_name",
        clear_rows={"start_row": 2, "end_row": 3},
    )
    assert (tmp_path / "out.xls").exists()


class _FakeWorksheet:
    def __init__(self, max_column: int = 3):
        self.max_column = max_column
        self.values = {}

    def cell(self, row: int, column: int, value=...):
        if value is not ...:
            self.values[(row, column)] = value
        return SimpleNamespace(value=self.values.get((row, column)))


class _FakeXlsSheet:
    def __init__(self):
        self.values = {}

    def write(self, row: int, col: int, value):
        self.values[(row, col)] = value


def test_process_data_to_rows_warning_branches(monkeypatch):
    writer = ExcelWriter()

    def fake_resolve(column, *_args, **_kwargs):
        if column in {"BAD_FIELD", "BAD_FIXED", "BAD_AUTO", "BAD_MONTH"}:
            raise ValueError("bad")
        return 1

    monkeypatch.setattr(writer, "_resolve_column_index_by_mode", fake_resolve)

    rows = writer._process_data_to_rows(
        data=[{"姓名": "张三"}],
        field_mappings={
            "姓名": {"source_column": "姓名", "target_column": "A"},
            "旧格式列": "BAD_FIELD",
        },
        headers={},
        max_columns=2,
        mapping_mode="column_name",
        fixed_values={"BAD_FIXED": "固定"},
        auto_number={"enabled": True, "column": "BAD_AUTO", "start_from": 1},
        month_type_mapping={"enabled": True, "target_column": "BAD_MONTH"},
        month_param="1",
    )

    assert rows == [["张三", ""]]


def test_write_data_to_worksheet_amount_and_warning_branches(monkeypatch):
    writer = ExcelWriter()
    ws = _FakeWorksheet(max_column=3)

    def fake_resolve(column, *_args, **_kwargs):
        if column in {"BAD_FIELD", "BAD_FIXED", "BAD_AUTO", "BAD_MONTH"}:
            raise ValueError("bad")
        if column == "金额":
            return 2
        if column == "姓名":
            return 1
        return 1

    monkeypatch.setattr(writer, "_resolve_column_index_by_mode", fake_resolve)

    writer._write_data_to_worksheet(
        ws=ws,
        data=[{"姓名": "张三", "金额": "12.5"}, {"姓名": "李四", "金额": "bad"}],
        field_mappings={
            "姓名": {"source_column": "姓名", "target_column": "姓名"},
            "金额": {"source_column": "金额", "target_column": "金额", "transform": "amount_decimal"},
            "旧格式列": "BAD_FIELD",
        },
        headers={},
        start_row=2,
        mapping_mode="column_name",
        fixed_values={"BAD_FIXED": "固定"},
        auto_number={"enabled": True, "column": "BAD_AUTO", "start_from": None},
        month_type_mapping={"enabled": True, "target_column": "BAD_MONTH"},
        month_param="1",
    )

    assert ws.values[(2, 2)] == 12.5
    assert ws.values[(3, 2)] == "bad"


def test_write_data_to_xls_sheet_warning_branches(monkeypatch):
    writer = ExcelWriter()
    ws = _FakeXlsSheet()

    def fake_resolve(column, *_args, **_kwargs):
        if column in {"BAD_FIELD", "BAD_FIXED", "BAD_AUTO", "BAD_MONTH"}:
            raise ValueError("bad")
        if column == "A":
            return 1
        return 1

    monkeypatch.setattr(writer, "_resolve_column_index_by_mode", fake_resolve)

    writer._write_data_to_xls_sheet(
        ws=ws,
        data=[{"姓名": "张三"}],
        field_mappings={
            "姓名": {"source_column": "姓名", "target_column": "A"},
            "旧格式列": "BAD_FIELD",
        },
        headers={},
        start_row=2,
        max_columns=2,
        mapping_mode="column_name",
        fixed_values={"BAD_FIXED": "固定"},
        auto_number={"enabled": True, "column": "BAD_AUTO", "start_from": 1},
        month_type_mapping={"enabled": True, "target_column": "BAD_MONTH"},
        month_param="1",
    )

    assert ws.values[(1, 0)] == "张三"


def test_calculate_month_value_bonus_compensation_and_out_of_range():
    writer = ExcelWriter()
    mapping = {"bonus_value": "奖金", "compensation_value": "补偿"}
    assert writer._calculate_month_value("13", mapping) is None
    assert writer._calculate_month_value("年终奖", mapping) == "奖金"
    assert writer._calculate_month_value("补偿金", mapping) == "补偿"
