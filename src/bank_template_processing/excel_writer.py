"""Excel 写入器模块。"""

from __future__ import annotations

import csv
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Mapping, Optional

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    openpyxl = None  # type: ignore
    InvalidFileException = type("InvalidFileException", (Exception,), {})  # type: ignore

try:
    import xlrd
except ImportError:
    xlrd = None  # type: ignore

try:
    import xlwt
except ImportError:
    xlwt = None  # type: ignore

try:
    from xlutils.copy import copy as xl_copy
except ImportError:
    xl_copy = None  # type: ignore

from .config_loader import ConfigError
from .sheet_utils import (
    column_letter_to_index,
    extract_headers_from_values,
    resolve_column_index,
    resolve_column_index_by_mode,
)


logger = logging.getLogger(__name__)


class ExcelError(Exception):
    """Excel 操作失败异常。"""

    pass


@dataclass(frozen=True)
class _CellProjection:
    """单个输出单元格投影。"""

    value: Any
    transform_type: str = "none"


class ExcelWriter:
    """Excel 写入器，支持 `.xlsx/.csv/.xls`。"""

    def __init__(self):
        logger.debug("ExcelWriter 初始化")

    def write_excel(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
        clear_rows: Mapping[str, Any] | None = None,
    ) -> None:
        """将数据写入 Excel 文件。"""
        logger.info(f"开始写入Excel文件: {template_path} -> {output_path}")
        logger.info(f"数据行数: {len(data)}")

        if start_row <= header_row:
            error_msg = f"配置错误: start_row ({start_row}) 必须大于 header_row ({header_row})"
            logger.error(error_msg)
            raise ConfigError(error_msg)

        if bank_branch_mapping and bank_branch_mapping.get("enabled"):
            logger.warning("配置警告: 'bank_branch_mapping' 已废弃，请使用 'field_mappings' 进行配置。")

        ext = Path(template_path).suffix.lower()
        try:
            if ext == ".xlsx":
                self._write_xlsx(
                    template_path,
                    data,
                    field_mappings,
                    output_path,
                    header_row,
                    start_row,
                    mapping_mode,
                    fixed_values,
                    auto_number,
                    bank_branch_mapping,
                    month_type_mapping,
                    month_param,
                    clear_rows,
                )
            elif ext == ".csv":
                self._write_csv(
                    template_path,
                    data,
                    field_mappings,
                    output_path,
                    header_row,
                    start_row,
                    mapping_mode,
                    fixed_values,
                    auto_number,
                    bank_branch_mapping,
                    month_type_mapping,
                    month_param,
                    clear_rows,
                )
            elif ext == ".xls":
                self._write_xls(
                    template_path,
                    data,
                    field_mappings,
                    output_path,
                    header_row,
                    start_row,
                    mapping_mode,
                    fixed_values,
                    auto_number,
                    bank_branch_mapping,
                    month_type_mapping,
                    month_param,
                    clear_rows,
                )
            else:
                raise ExcelError(f"不支持的文件格式: {ext}")
        except Exception as exc:
            if isinstance(exc, (ConfigError, ExcelError, FileNotFoundError)):
                raise
            error_msg = f"写入文件失败: {exc}"
            logger.error(error_msg, exc_info=True)
            raise ExcelError(error_msg) from exc

        logger.info(f"文件写入成功: {output_path}")

    def _write_xlsx(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
        clear_rows: Mapping[str, Any] | None = None,
    ) -> None:
        """使用 openpyxl 写入 `.xlsx` 文件。"""
        del bank_branch_mapping
        logger.debug(f"使用openpyxl写入xlsx文件: {template_path}")

        if openpyxl is None:
            raise ExcelError("openpyxl未安装，无法处理.xlsx文件")

        try:
            workbook = openpyxl.load_workbook(template_path)
        except InvalidFileException as exc:
            raise ExcelError(f"无效的Excel文件: {exc}") from exc

        worksheet = workbook.worksheets[0] if workbook.worksheets else workbook.active
        if worksheet is None:
            raise ExcelError("模板文件没有工作表")

        headers = self._extract_headers_from_xlsx(worksheet, header_row)
        self._clear_xlsx_rows(worksheet, data, start_row, clear_rows)
        self._write_data_to_worksheet(
            worksheet,
            data,
            field_mappings,
            headers,
            start_row,
            mapping_mode,
            fixed_values,
            auto_number,
            None,
            month_type_mapping,
            month_param,
        )
        workbook.save(output_path)
        logger.debug(f"xlsx文件已保存: {output_path}")

    def _write_csv(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
        clear_rows: Mapping[str, Any] | None = None,
    ) -> None:
        """使用 csv 模块写入 `.csv` 文件。"""
        del bank_branch_mapping
        logger.debug(f"使用csv模块写入csv文件: {template_path}")

        with open(template_path, "r", encoding="utf-8-sig") as file:
            reader = csv.reader(file)
            rows = list(reader)

        if header_row > 0:
            if len(rows) < header_row:
                raise ExcelError(f"模板文件行数不足，无法读取表头行: {header_row}")
            header_values = rows[header_row - 1]
            headers = extract_headers_from_values(header_values)
            max_columns = len(header_values)
            logger.debug(f"读取到 {len(headers)} 个表头字段")
        else:
            headers = {}
            max_columns = len(rows[0]) if rows else 0
            logger.debug("header_row = 0，跳过读取表头（使用列标识符）")

        data_rows = self._process_data_to_rows(
            data,
            field_mappings,
            headers,
            max_columns,
            mapping_mode,
            fixed_values,
            auto_number,
            None,
            month_type_mapping,
            month_param,
        )

        clear_config = clear_rows or {}
        clear_end = clear_config.get("end_row", clear_config.get("data_end_row"))
        if clear_end is not None:
            clear_start = clear_config.get("start_row", start_row)
            if clear_start > clear_end:
                raise ExcelError("clear_rows.start_row 不能大于 end_row")
            clear_count = clear_end - clear_start + 1
            rows_before = rows[: clear_start - 1]
            if len(rows_before) < clear_start - 1:
                rows_before = rows_before + [[""] * max_columns for _ in range(clear_start - 1 - len(rows_before))]
            rows_after = rows[clear_end:] if clear_end < len(rows) else []
            filler_rows = [[""] * max_columns for _ in range(max(0, clear_count - len(data_rows)))]
            output_rows = rows_before + data_rows + filler_rows + rows_after
        else:
            output_rows = rows[: start_row - 1]
            if len(output_rows) < start_row - 1:
                output_rows = output_rows + [[""] * max_columns for _ in range(start_row - 1 - len(output_rows))]
            output_rows.extend(data_rows)

        with open(output_path, "w", encoding="utf-8-sig", newline="") as file:
            writer = csv.writer(file)
            writer.writerows(output_rows)
        logger.debug(f"csv文件已保存: {output_path}")

    def _write_xls(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
        clear_rows: Mapping[str, Any] | None = None,
    ) -> None:
        """使用 xlwt 写入 `.xls` 文件。"""
        del bank_branch_mapping
        logger.debug(f"使用xlwt写入xls文件: {template_path}")

        if xlwt is None:
            raise ExcelError("xlwt未安装，无法处理.xls文件")
        if xlrd is None:
            raise ExcelError("xlrd未安装，无法读取.xls模板文件")
        if xl_copy is None:
            raise ExcelError("xlutils未安装，无法保留.xls模板格式")

        try:
            workbook_template = xlrd.open_workbook(template_path, formatting_info=True)
        except Exception as exc:
            raise ExcelError(f"无法读取Excel文件: {exc}") from exc

        workbook_output = xl_copy(workbook_template)
        worksheet_template = workbook_template.sheet_by_index(0)
        worksheet_output = workbook_output.get_sheet(0)

        if header_row > 0:
            header_values = [worksheet_template.cell_value(header_row - 1, idx) for idx in range(worksheet_template.ncols)]
            headers = extract_headers_from_values(header_values)
            logger.debug(f"读取到 {len(headers)} 个表头字段")
        else:
            headers = {}
            logger.debug("header_row = 0，跳过读取表头（使用列标识符）")

        clear_config = clear_rows or {}
        clear_end = clear_config.get("end_row", clear_config.get("data_end_row"))
        if clear_end is not None:
            clear_start = clear_config.get("start_row", start_row)
            if clear_start > clear_end:
                raise ConfigError("clear_rows.start_row 不能大于 end_row")
            clear_count = clear_end - clear_start + 1
            if len(data) > clear_count:
                raise ConfigError("clear_rows 范围不足以容纳全部数据，请增大 end_row")
            logger.debug(f"清理数据区：{clear_start}-{clear_end} (覆盖为空)")
            for row_idx in range(clear_start - 1, clear_end):
                for col_idx in range(worksheet_template.ncols):
                    worksheet_output.write(row_idx, col_idx, "")
        elif worksheet_template.nrows > start_row - 1:
            logger.debug(f"清除从第 {start_row} 行开始的数据 (覆盖为空)")
            for row_idx in range(start_row - 1, worksheet_template.nrows):
                for col_idx in range(worksheet_template.ncols):
                    worksheet_output.write(row_idx, col_idx, "")

        self._write_data_to_xls_sheet(
            worksheet_output,
            data,
            field_mappings,
            headers,
            start_row,
            worksheet_template.ncols,
            mapping_mode,
            fixed_values,
            auto_number,
            None,
            month_type_mapping,
            month_param,
        )
        workbook_output.save(output_path)
        logger.debug(f"xls文件已保存: {output_path}")

    def _process_data_to_rows(
        self,
        data: list,
        field_mappings: dict,
        headers: dict,
        max_columns: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
    ) -> list[list[str]]:
        """处理数据并转换为 CSV 行。"""
        projections = self._project_rows(
            data,
            field_mappings,
            headers,
            max_columns,
            mapping_mode,
            fixed_values,
            auto_number,
            bank_branch_mapping,
            month_type_mapping,
            month_param,
            bounded=True,
        )
        rows: list[list[str]] = []
        for projection in projections:
            row_output = [""] * max_columns
            for col_idx, cell in projection.items():
                if 1 <= col_idx <= max_columns:
                    row_output[col_idx - 1] = "" if cell.value is None else str(cell.value)
            rows.append(row_output)
        return rows

    def _write_data_to_worksheet(
        self,
        ws,
        data: list,
        field_mappings: dict,
        headers: dict,
        start_row: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
    ) -> None:
        """将数据写入 openpyxl 工作表。"""
        projections = self._project_rows(
            data,
            field_mappings,
            headers,
            ws.max_column,
            mapping_mode,
            fixed_values,
            auto_number,
            bank_branch_mapping,
            month_type_mapping,
            month_param,
            bounded=False,
        )
        for row_idx, projection in enumerate(projections, start=start_row):
            for col_idx, cell in projection.items():
                ws.cell(row_idx, col_idx, self._coerce_xlsx_value(cell))
        logger.debug(f"已写入 {len(data)} 行数据到工作表")

    def _write_data_to_xls_sheet(
        self,
        ws,
        data: list,
        field_mappings: dict,
        headers: dict,
        start_row: int,
        max_columns: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
    ) -> None:
        """将数据写入 xlwt 工作表。"""
        projections = self._project_rows(
            data,
            field_mappings,
            headers,
            max_columns,
            mapping_mode,
            fixed_values,
            auto_number,
            bank_branch_mapping,
            month_type_mapping,
            month_param,
            bounded=True,
        )
        for row_idx, projection in enumerate(projections, start=start_row):
            for col_idx, cell in projection.items():
                if 1 <= col_idx <= max_columns:
                    ws.write(row_idx - 1, col_idx - 1, cell.value)
        logger.debug(f"已写入 {len(data)} 行数据到xls工作表")

    def _calculate_month_value(
        self,
        month_param: Optional[str],
        month_type_mapping: Mapping[str, Any],
    ) -> Optional[str]:
        """计算 month_type_mapping 的统一填充值。"""
        if month_param is None:
            return None

        try:
            month_num = int(month_param)
            if 1 <= month_num <= 12:
                month_format = month_type_mapping.get("month_format", "{month}月收入")
                try:
                    return month_format.format(month=f"{month_num:02d}")
                except KeyError as exc:
                    raise ConfigError(f"month_format 缺少变量: {exc}") from exc
                except Exception as exc:
                    raise ConfigError(f"month_format 格式错误: {exc}") from exc
        except ValueError:
            pass

        if month_param == "年终奖":
            return month_type_mapping.get("bonus_value", "年终奖")
        if month_param == "补偿金":
            return month_type_mapping.get("compensation_value", "补偿金")
        return None

    def _column_letter_to_index(self, column: str) -> int:
        """包装共享列字母解析。"""
        return column_letter_to_index(column)

    def _resolve_column_index(
        self,
        column_spec,
        headers: Optional[dict] = None,
        max_columns: Optional[int] = None,
        strict_bounds: bool = False,
    ) -> int:
        """包装共享列解析。"""
        return resolve_column_index(column_spec, headers=headers, max_columns=max_columns, strict_bounds=strict_bounds)

    def _resolve_column_index_by_mode(
        self,
        column_spec,
        headers: Optional[dict],
        max_columns: Optional[int],
        mapping_mode: str,
    ) -> int:
        """包装共享按模式列解析。"""
        return resolve_column_index_by_mode(
            column_spec,
            headers=headers,
            max_columns=max_columns,
            mapping_mode=mapping_mode,
            logger_instance=logger,
        )

    def _resolve_required_column(
        self,
        column_spec,
        headers: Optional[dict],
        max_columns: Optional[int],
        mapping_mode: str,
        error_label: str,
    ) -> int | None:
        try:
            return self._resolve_column_index_by_mode(column_spec, headers, max_columns, mapping_mode)
        except ValueError as exc:
            if mapping_mode == "column_name":
                raise ConfigError(f"{error_label}: {exc}") from exc
            logger.warning(f"{error_label}: {exc}")
            return None

    def _extract_headers_from_xlsx(self, worksheet, header_row: int) -> dict[str, int]:
        if header_row <= 0:
            logger.debug("header_row = 0，跳过读取表头（使用列标识符）")
            return {}
        header_values = [worksheet.cell(header_row, col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
        headers = extract_headers_from_values(header_values)
        logger.debug(f"读取到 {len(headers)} 个表头字段")
        return headers

    def _clear_xlsx_rows(
        self,
        worksheet,
        data: list,
        start_row: int,
        clear_rows: Mapping[str, Any] | None,
    ) -> None:
        clear_config = clear_rows or {}
        clear_end = clear_config.get("end_row", clear_config.get("data_end_row"))
        if clear_end is not None:
            clear_start = clear_config.get("start_row", start_row)
            if clear_start > clear_end:
                raise ConfigError("clear_rows.start_row 不能大于 end_row")
            clear_count = clear_end - clear_start + 1
            if len(data) > clear_count:
                worksheet.insert_rows(clear_end + 1, amount=len(data) - clear_count)
            for row_idx in range(clear_start, clear_end + 1):
                for col_idx in range(1, worksheet.max_column + 1):
                    worksheet.cell(row_idx, col_idx, None)
            logger.debug(f"清理数据区：{clear_start}-{clear_end}")
            return

        logger.debug(f"清除从第 {start_row} 行开始的数据")
        if worksheet.max_row >= start_row:
            worksheet.delete_rows(start_row, worksheet.max_row - start_row + 1)

    def _project_rows(
        self,
        data: list,
        field_mappings: dict,
        headers: dict,
        max_columns: int,
        mapping_mode: str,
        fixed_values: Mapping[str, Any] | None = None,
        auto_number: Mapping[str, Any] | None = None,
        bank_branch_mapping: Mapping[str, Any] | None = None,
        month_type_mapping: Mapping[str, Any] | None = None,
        month_param: Optional[str] = None,
        *,
        bounded: bool,
    ) -> list[dict[int, _CellProjection]]:
        del bank_branch_mapping
        projections: list[dict[int, _CellProjection]] = []

        current_number = None
        if auto_number and auto_number.get("enabled"):
            current_number = auto_number.get("start_from", 1)

        month_value = None
        if month_type_mapping and month_type_mapping.get("enabled"):
            month_value = self._calculate_month_value(month_param, month_type_mapping)

        for row_data in data:
            row_projection: dict[int, _CellProjection] = {}

            for template_column, mapping_config in field_mappings.items():
                source_column, target_column, transform_type = self._normalize_field_mapping(template_column, mapping_config)
                value = row_data.get(source_column, "")
                col_idx = self._resolve_required_column(
                    target_column,
                    headers,
                    max_columns,
                    mapping_mode,
                    f"无法解析字段 '{template_column}' 的目标列 '{target_column}'",
                )
                if col_idx is None:
                    continue
                self._set_projection_value(
                    row_projection,
                    col_idx,
                    _CellProjection(value=value, transform_type=transform_type),
                    max_columns=max_columns,
                    bounded=bounded,
                )

            if fixed_values:
                for column, value in fixed_values.items():
                    col_idx = self._resolve_required_column(
                        column,
                        headers,
                        max_columns,
                        mapping_mode,
                        f"无法解析固定值列 '{column}'",
                    )
                    if col_idx is None:
                        continue
                    self._set_projection_value(
                        row_projection,
                        col_idx,
                        _CellProjection(value=value),
                        max_columns=max_columns,
                        bounded=bounded,
                    )

            if auto_number and auto_number.get("enabled"):
                column = auto_number.get("column", auto_number.get("column_name", "A"))
                col_idx = self._resolve_required_column(
                    column,
                    headers,
                    max_columns,
                    mapping_mode,
                    f"无法解析自动编号列 '{column}'",
                )
                if col_idx is None:
                    continue
                self._set_projection_value(
                    row_projection,
                    col_idx,
                    _CellProjection(value=current_number),
                    max_columns=max_columns,
                    bounded=bounded,
                )
                if current_number is not None:
                    current_number += 1

            if month_value is not None and month_type_mapping:
                target_column = month_type_mapping.get("target_column", "C")
                col_idx = self._resolve_required_column(
                    target_column,
                    headers,
                    max_columns,
                    mapping_mode,
                    f"无法解析月类型映射列 '{target_column}'",
                )
                if col_idx is None:
                    continue
                self._set_projection_value(
                    row_projection,
                    col_idx,
                    _CellProjection(value=month_value),
                    max_columns=max_columns,
                    bounded=bounded,
                )

            projections.append(row_projection)

        return projections

    def _normalize_field_mapping(self, template_column: str, mapping_config: Any) -> tuple[str, Any, str]:
        if isinstance(mapping_config, dict):
            source_column = mapping_config.get("source_column")
            target_column = mapping_config.get("target_column", template_column)
            transform_type = mapping_config.get("transform", "none")
            return source_column, target_column, transform_type
        return template_column, mapping_config, "none"

    def _set_projection_value(
        self,
        row_projection: dict[int, _CellProjection],
        col_idx: int,
        cell: _CellProjection,
        *,
        max_columns: int,
        bounded: bool,
    ) -> None:
        if col_idx < 1:
            return
        if bounded and col_idx > max_columns:
            return
        row_projection[col_idx] = cell

    def _coerce_xlsx_value(self, cell: _CellProjection) -> Any:
        if cell.transform_type == "amount_decimal" and cell.value is not None:
            try:
                if isinstance(cell.value, str):
                    return float(cell.value)
                return cell.value
            except (ValueError, TypeError):
                return cell.value
        return cell.value
