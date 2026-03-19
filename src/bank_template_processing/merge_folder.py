"""批量合并目录模式实现。"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Mapping

import openpyxl
import xlrd

from .config_loader import ConfigError, get_unit_config
from .config_types import RuleGroupConfig
from .pipeline import (
    ProcessingContext,
    calculate_stats as pipeline_calculate_stats,
    enrich_error_context,
    needs_transformations as pipeline_needs_transformations,
    split_validation_rules as pipeline_split_validation_rules,
    transform_rows as pipeline_transform_rows,
    validate_rows as pipeline_validate_rows,
)
from .excel_writer import ExcelWriter
from .sheet_utils import (
    convert_xls_cell,
    extract_headers,
    get_cell_value,
    is_empty_value,
    resolve_column_index_by_mode,
)
from .validator import ValidationError


MERGE_FILE_PATTERN = re.compile(r"^(?P<prefix>.+)_(?P<count>\d+)人_金额(?P<amount>-?\d+(?:\.\d+)?)元$")
MERGE_COUNT_PATTERN = re.compile(r"(?P<count>\d+)人")
MERGE_AMOUNT_PATTERN = re.compile(r"金额(?P<amount>-?\d+(?:\.\d+)?)元")
SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}
MERGE_MONTH_SOURCE_COLUMN = "__merge_month_value__"
MERGE_SOURCE_FILE_COLUMN = "__merge_source_file__"


logger = logging.getLogger(__name__)


class MergeFolderError(Exception):
    """批量合并模式错误。"""

    pass


@dataclass(frozen=True)
class MergeInputFile:
    """待合并输入文件元信息。"""

    path: Path
    unit_name: str
    template_name: str
    count: int | None
    amount: float | None


@dataclass
class MergeTask:
    """单个汇总文件的执行任务。"""

    unit_name: str
    template_name: str
    rule_group: str
    group_config: RuleGroupConfig | dict[str, Any]
    template_path: str
    group_data: list[dict]
    month_param: str
    count: int
    amount: float


def parse_merge_filename(
    file_path: Path,
    unit_names: list[str],
    template_names_by_unit: Mapping[str, set[str]] | None = None,
) -> MergeInputFile:
    """解析合并输入文件名并提取元信息。"""
    if file_path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise MergeFolderError(f"不支持的合并文件格式: {file_path.name}")

    match = MERGE_FILE_PATTERN.match(file_path.stem)
    if not match:
        if template_names_by_unit is None:
            raise MergeFolderError(f"文件名不符合合并命名规则: {file_path.name}")
        return _infer_merge_input_file_from_name(file_path, unit_names, template_names_by_unit)

    prefix = match.group("prefix")
    count = int(match.group("count"))
    amount = float(match.group("amount"))

    try:
        unit_name, template_name = _split_prefix_to_unit_and_template(prefix, unit_names)
    except MergeFolderError:
        if template_names_by_unit is None:
            raise
        return _infer_merge_input_file_from_name(
            file_path,
            unit_names,
            template_names_by_unit,
            parsed_count=count,
            parsed_amount=amount,
        )

    return MergeInputFile(path=file_path, unit_name=unit_name, template_name=template_name, count=count, amount=amount)


def resolve_rule_group_for_template(
    config: Mapping[str, Any],
    unit_name: str,
    template_name: str,
) -> tuple[str, RuleGroupConfig]:
    """
    根据“单位+模板名”解析规则组。

    优先级：
    1. template_selector.default_group_name/special_group_name
    2. 规则组 template_path 的 stem
    """
    org_units = config.get("organization_units", {})
    if unit_name not in org_units:
        raise ConfigError(f"配置文件中未找到单位配置：{unit_name}")

    unit_config = org_units[unit_name]

    # 第一步：按组名匹配（优先）
    selector_config = unit_config.get("template_selector", {}) if isinstance(unit_config, dict) else {}
    default_group_name = selector_config.get("default_group_name", "default")
    special_group_name = selector_config.get("special_group_name", "special")

    group_name_candidates: list[str] = []
    if template_name == default_group_name:
        group_name_candidates.append("default")
    if template_name == special_group_name:
        group_name_candidates.append("crossbank")

    if group_name_candidates:
        unique_candidates = sorted(set(group_name_candidates))
        if len(unique_candidates) != 1:
            raise MergeFolderError(f"模板名称 '{template_name}' 同时命中多个组名配置: {', '.join(unique_candidates)}")
        rule_group = unique_candidates[0]
        return rule_group, get_unit_config(config, unit_name, rule_group)

    # 第二步：按模板文件名 stem 回退匹配
    stem_candidates: list[str] = []
    if isinstance(unit_config, dict) and "default" in unit_config:
        for rule_name, rule_config in unit_config.items():
            if rule_name == "template_selector":
                continue
            if not isinstance(rule_config, dict):
                continue
            template_path = rule_config.get("template_path")
            if isinstance(template_path, str) and Path(template_path).stem == template_name:
                stem_candidates.append(rule_name)
    else:
        template_path = unit_config.get("template_path") if isinstance(unit_config, dict) else None
        if isinstance(template_path, str) and Path(template_path).stem == template_name:
            stem_candidates.append("default")

    unique_stem_candidates = sorted(set(stem_candidates))
    if not unique_stem_candidates:
        raise MergeFolderError(
            f"无法为单位 '{unit_name}' 的模板名称 '{template_name}' 匹配规则组，请检查组名或模板路径配置"
        )
    if len(unique_stem_candidates) > 1:
        raise MergeFolderError(
            f"单位 '{unit_name}' 的模板名称 '{template_name}' 匹配到多个规则组: {', '.join(unique_stem_candidates)}"
        )

    rule_group = unique_stem_candidates[0]
    return rule_group, get_unit_config(config, unit_name, rule_group)


def infer_month_param_from_values(
    month_values: set[str],
    month_type_mapping: dict,
    allow_conflict: bool = False,
    logger: Any | None = None,
) -> str:
    """根据文件中月类型列值推断 month 参数。"""
    if not month_values:
        raise MergeFolderError("month_type_mapping 已启用，但未在输入文件中读取到月类型值")

    inferred_params = {_infer_month_param_from_single_value(value, month_type_mapping) for value in month_values}
    if len(inferred_params) != 1:
        if allow_conflict:
            selected_param = _select_month_param_on_conflict(inferred_params)
            if logger:
                logger.warning(
                    "检测到同一分组存在冲突的月类型值 %s，已忽略差异并使用 '%s' 继续合并",
                    sorted(month_values),
                    selected_param,
                )
            return selected_param
        raise MergeFolderError(
            f"同一分组存在冲突的月类型值: {sorted(month_values)}，推断结果: {sorted(inferred_params)}"
        )
    return inferred_params.pop()


def prepare_merge_tasks(
    merge_folder_path: str,
    config: Mapping[str, Any],
    resolve_path_fn: Callable[[str], str],
    apply_transformations_fn: Callable[..., list] | None,
    needs_transformations_fn: Callable[..., bool] | None,
    calculate_stats_fn: Callable[..., tuple[int, float]] | None,
    needs_month_for_filename: bool,
    logger: Any,
) -> list[MergeTask]:
    """扫描并准备批量合并任务。"""
    transform_fn = apply_transformations_fn or pipeline_transform_rows
    needs_transform_fn = needs_transformations_fn or pipeline_needs_transformations
    stats_fn = calculate_stats_fn or pipeline_calculate_stats

    merge_folder = Path(merge_folder_path)
    if not merge_folder.exists():
        raise FileNotFoundError(f"合并目录不存在: {merge_folder}")
    if not merge_folder.is_dir():
        raise MergeFolderError(f"合并路径不是目录: {merge_folder}")

    organization_units = config.get("organization_units", {})
    if not isinstance(organization_units, dict) or not organization_units:
        raise ConfigError("配置缺少 organization_units，无法执行批量合并")

    unit_names = list(organization_units.keys())
    template_names_by_unit = _build_template_name_candidates_by_unit(config)
    input_files = _scan_merge_input_files(merge_folder, unit_names, template_names_by_unit)
    logger.info(
        "批量合并扫描完成：目录 %s 共发现 %s 个输入文件（排序规则=mtime asc, tie=name asc）",
        merge_folder,
        len(input_files),
    )

    grouped_files: dict[tuple[str, str], list[MergeInputFile]] = {}
    for input_file in input_files:
        group_key = (input_file.unit_name, input_file.template_name)
        grouped_files.setdefault(group_key, []).append(input_file)

    logger.info("按单位+模板分组完成：共 %s 组", len(grouped_files))
    merge_tasks: list[MergeTask] = []

    for unit_name, template_name in sorted(grouped_files.keys()):
        files = grouped_files[(unit_name, template_name)]
        logger.info("处理合并分组：%s_%s（%s 个文件）", unit_name, template_name, len(files))

        rule_group, group_config = resolve_rule_group_for_template(config, unit_name, template_name)
        context = ProcessingContext(unit_name=unit_name, rule_group=rule_group, template_name=template_name)
        template_path_raw = group_config.get("template_path", "")
        if not template_path_raw:
            error = ConfigError(f"单位 '{unit_name}' 的规则组 '{rule_group}' 未配置 template_path")
            raise enrich_error_context(error, "批量合并配置解析", context) from error
        template_path = resolve_path_fn(template_path_raw)

        merged_group_data: list[dict] = []
        merged_month_values: set[str] = set()
        count_from_name = 0
        amount_from_name = 0.0
        has_complete_name_count = True
        has_complete_name_amount = True

        for file_meta in files:
            file_context = context.with_source_file(file_meta.path.name)
            try:
                file_rows, month_values = _read_generated_file_rows(file_meta.path, group_config)
            except MergeFolderError as exc:
                raise enrich_error_context(exc, "批量合并读取", file_context) from exc
            merged_group_data.extend(file_rows)
            merged_month_values.update(month_values)
            logger.info(
                "已读取文件 %s：提取 %s 行，文件名统计 人数=%s 金额=%.2f",
                file_meta.path.name,
                len(file_rows),
                file_meta.count if file_meta.count is not None else "未提供",
                file_meta.amount if file_meta.amount is not None else 0.0,
            )

            if file_meta.count is None:
                has_complete_name_count = False
            else:
                count_from_name += file_meta.count

            if file_meta.amount is None:
                has_complete_name_amount = False
            else:
                amount_from_name += file_meta.amount

        validation_rules = group_config.get("validation_rules", {})
        pre_transform_rules, post_transform_rules = pipeline_split_validation_rules(validation_rules)
        if pre_transform_rules:
            logger.info("分组 %s_%s 开始必填校验", unit_name, template_name)
            pipeline_validate_rows(
                merged_group_data,
                pre_transform_rules,
                context=context,
                source_file_field=MERGE_SOURCE_FILE_COLUMN,
            )

        field_mappings = group_config.get("field_mappings", {})
        transformations = group_config.get("transformations", {})
        if needs_transform_fn(field_mappings):
            logger.info("分组 %s_%s 开始数据转换", unit_name, template_name)
            merged_group_data = transform_fn(
                merged_group_data,
                transformations,
                field_mappings,
                context=context,
                source_file_field=MERGE_SOURCE_FILE_COLUMN,
            )
        if post_transform_rules:
            logger.info("分组 %s_%s 开始类型/范围校验", unit_name, template_name)
            pipeline_validate_rows(
                merged_group_data,
                post_transform_rules,
                context=context,
                source_file_field=MERGE_SOURCE_FILE_COLUMN,
            )
        try:
            count_from_data, amount_from_data = stats_fn(merged_group_data, field_mappings, transformations)
        except ValidationError as exc:
            raise enrich_error_context(exc, "批量合并统计", context) from exc

        if has_complete_name_count and count_from_data != count_from_name:
            error = MergeFolderError(
                f"分组 '{unit_name}_{template_name}' 人数校验失败：文件名累加={count_from_name}，数据重算={count_from_data}"
            )
            raise enrich_error_context(error, "批量合并统计校验", context) from error
        if not has_complete_name_count:
            logger.info("分组 %s_%s 跳过文件名人数校验：输入文件名未提供完整人数信息", unit_name, template_name)

        if has_complete_name_amount and abs(amount_from_data - amount_from_name) > 0.01:
            error = MergeFolderError(
                "分组 '{0}_{1}' 金额校验失败：文件名累加={2:.2f}，数据重算={3:.2f}".format(
                    unit_name,
                    template_name,
                    amount_from_name,
                    amount_from_data,
                )
            )
            raise enrich_error_context(error, "批量合并统计校验", context) from error
        if not has_complete_name_amount:
            logger.info("分组 %s_%s 跳过文件名金额校验：输入文件名未提供完整金额信息", unit_name, template_name)

        month_param = ""
        month_type_mapping = group_config.get("month_type_mapping", {})
        output_group_config = group_config
        if isinstance(month_type_mapping, dict) and month_type_mapping.get("enabled"):
            output_group_config = _build_merge_output_group_config(group_config, keep_row_month_values=True)
            if needs_month_for_filename:
                try:
                    month_param = infer_month_param_from_values(
                        merged_month_values,
                        month_type_mapping,
                        allow_conflict=True,
                        logger=logger,
                    )
                except MergeFolderError as exc:
                    raise enrich_error_context(exc, "批量合并月份推断", context) from exc
                logger.info("分组 %s_%s 月份参数推断成功：%s", unit_name, template_name, month_param)
            else:
                logger.info("分组 %s_%s 跳过月份参数推断：输出文件名模板未使用 {month}", unit_name, template_name)

        merge_tasks.append(
            MergeTask(
                unit_name=unit_name,
                template_name=template_name,
                rule_group=rule_group,
                group_config=output_group_config,
                template_path=template_path,
                group_data=merged_group_data,
                month_param=month_param,
                count=count_from_data,
                amount=amount_from_data,
            )
        )

    return merge_tasks


def _scan_merge_input_files(
    merge_folder: Path,
    unit_names: list[str],
    template_names_by_unit: Mapping[str, set[str]] | None = None,
) -> list[MergeInputFile]:
    excel_files = [
        path for path in merge_folder.iterdir() if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    ]
    if not excel_files:
        raise MergeFolderError(f"目录中未找到可合并的 Excel 文件: {merge_folder}")
    excel_files.sort(key=_merge_input_file_sort_key)

    parsed_files: list[MergeInputFile] = []
    for file_path in excel_files:
        parsed_files.append(parse_merge_filename(file_path, unit_names, template_names_by_unit))
    return parsed_files


def _build_template_name_candidates_by_unit(
    config: Mapping[str, Any],
) -> dict[str, set[str]]:
    """按单位收集合并文件名中可识别的模板名候选。"""
    organization_units = config.get("organization_units", {})
    if not isinstance(organization_units, dict):
        return {}

    candidates_by_unit: dict[str, set[str]] = {}
    for unit_name, unit_config in organization_units.items():
        candidates: set[str] = set()
        if not isinstance(unit_config, dict):
            candidates_by_unit[unit_name] = candidates
            continue

        selector_config = unit_config.get("template_selector", {})
        if isinstance(selector_config, dict) and selector_config.get("enabled", False):
            candidates.add(str(selector_config.get("default_group_name", "default")))
            if "crossbank" in unit_config:
                candidates.add(str(selector_config.get("special_group_name", "special")))

        if "default" in unit_config:
            for rule_name, rule_config in unit_config.items():
                if rule_name in {"template_selector", "input_filename_routing"}:
                    continue
                if not isinstance(rule_config, dict):
                    continue
                template_path = rule_config.get("template_path")
                if isinstance(template_path, str) and template_path.strip():
                    candidates.add(Path(template_path).stem)
        else:
            template_path = unit_config.get("template_path")
            if isinstance(template_path, str) and template_path.strip():
                candidates.add(Path(template_path).stem)

        candidates_by_unit[unit_name] = {candidate for candidate in candidates if candidate}

    return candidates_by_unit


def _merge_input_file_sort_key(file_path: Path) -> tuple[float, str]:
    try:
        mtime = file_path.stat().st_mtime
    except OSError as e:
        raise MergeFolderError(f"读取文件修改时间失败: {file_path.name}: {e}") from e
    return mtime, file_path.name


def _split_prefix_to_unit_and_template(prefix: str, unit_names: list[str]) -> tuple[str, str]:
    for unit_name in sorted(unit_names, key=len, reverse=True):
        marker = f"{unit_name}_"
        if prefix.startswith(marker):
            template_name = prefix[len(marker) :]
            if not template_name:
                raise MergeFolderError(f"文件名前缀缺少模板名称: {prefix}")
            return str(unit_name), str(template_name)

    raise MergeFolderError(f"文件名前缀无法匹配单位名称: {prefix}")


def _infer_merge_input_file_from_name(
    file_path: Path,
    unit_names: list[str],
    template_names_by_unit: Mapping[str, set[str]],
    *,
    parsed_count: int | None = None,
    parsed_amount: float | None = None,
) -> MergeInputFile:
    """在默认命名规则之外，按配置中的单位/模板名称回推文件元信息。"""
    stem = file_path.stem
    unit_name = _select_unique_filename_candidate(stem, unit_names, "单位名称", file_path.name)
    template_candidates = template_names_by_unit.get(unit_name, set())
    if not template_candidates:
        raise MergeFolderError(f"文件名无法匹配模板名称: {file_path.name}")

    remainder = stem.replace(unit_name, " ", 1)
    template_name = _select_unique_filename_candidate(remainder, template_candidates, "模板名称", file_path.name)

    count = parsed_count if parsed_count is not None else _extract_count_from_name(stem)
    amount = parsed_amount if parsed_amount is not None else _extract_amount_from_name(stem)
    return MergeInputFile(
        path=file_path,
        unit_name=unit_name,
        template_name=template_name,
        count=count,
        amount=amount,
    )


def _select_unique_filename_candidate(
    text: str,
    candidates: list[str] | set[str],
    label: str,
    file_name: str,
) -> str:
    unique_matches: set[str] = set()
    for candidate in candidates:
        if isinstance(candidate, str) and candidate and candidate in text:
            unique_matches.add(candidate)

    matches: list[str] = sorted(unique_matches, key=lambda value: len(value), reverse=True)
    if not matches:
        raise MergeFolderError(f"文件名无法匹配{label}: {file_name}")

    primary = matches[0]
    conflicts: list[str] = []
    for candidate in matches[1:]:
        if candidate not in primary and primary not in candidate:
            conflicts.append(candidate)

    if conflicts:
        matched = "、".join([primary, *conflicts])
        raise MergeFolderError(f"文件名匹配到多个{label}: {file_name}: {matched}")
    return primary


def _extract_count_from_name(stem: str) -> int | None:
    match = MERGE_COUNT_PATTERN.search(stem)
    if match is None:
        return None
    return int(match.group("count"))


def _extract_amount_from_name(stem: str) -> float | None:
    match = MERGE_AMOUNT_PATTERN.search(stem)
    if match is None:
        return None
    return float(match.group("amount"))


def _read_generated_file_rows(
    file_path: Path,
    group_config: RuleGroupConfig | dict[str, Any],
) -> tuple[list[dict], set[str]]:
    rows = _read_all_rows(file_path)
    max_columns = max((len(row) for row in rows), default=0)

    header_row = group_config.get("header_row", 1)
    clear_rows = group_config.get("clear_rows", {})
    if clear_rows and not isinstance(clear_rows, dict):
        raise MergeFolderError(f"文件 {file_path.name} 的 clear_rows 配置无效: {clear_rows}")

    start_row = group_config.get("start_row", header_row + 1)
    if isinstance(clear_rows, dict):
        start_row = clear_rows.get("start_row", start_row)
    if not isinstance(start_row, int) or start_row < 1:
        raise MergeFolderError(f"文件 {file_path.name} 的 start_row 配置无效: {start_row}")

    end_row = len(rows)
    if isinstance(clear_rows, dict):
        clear_end = clear_rows.get("end_row", clear_rows.get("data_end_row"))
        if clear_end is not None:
            if not isinstance(clear_end, int) or clear_end < 1:
                raise MergeFolderError(f"文件 {file_path.name} 的 clear_rows.end_row 配置无效: {clear_end}")
            if clear_end < start_row:
                raise MergeFolderError(f"文件 {file_path.name} 的 clear_rows.end_row 不能小于 start_row")
            end_row = min(clear_end, len(rows))

    headers = _extract_headers(rows, header_row, file_path)
    field_mappings = group_config.get("field_mappings", {})

    bindings = _build_field_bindings(field_mappings, headers, max_columns, None, file_path)
    month_col_idx = _resolve_month_column(group_config, headers, max_columns, None, file_path)

    data_rows: list[dict] = []
    month_values: set[str] = set()

    for row_number in range(start_row, end_row + 1):
        row_values = rows[row_number - 1]
        row_dict: dict[str, Any] = {}
        has_data = False

        for source_column, col_idx in bindings:
            value = _get_cell_value(row_values, col_idx)
            row_dict[source_column] = value
            if not _is_empty_value(value):
                has_data = True

        if not has_data:
            continue

        if month_col_idx is not None:
            month_value = _get_cell_value(row_values, month_col_idx)
            if not _is_empty_value(month_value):
                month_values.add(str(month_value).strip())
            row_dict[MERGE_MONTH_SOURCE_COLUMN] = month_value
        row_dict[MERGE_SOURCE_FILE_COLUMN] = file_path.name

        data_rows.append(row_dict)

    return data_rows, month_values


def _build_field_bindings(
    field_mappings: dict,
    headers: dict[str, int],
    max_columns: int,
    writer: ExcelWriter | None,
    file_path: Path,
) -> list[tuple[str, int]]:
    del writer
    bindings: list[tuple[str, int]] = []
    for template_column, mapping_config in field_mappings.items():
        if isinstance(mapping_config, dict):
            source_column = mapping_config.get("source_column")
            target_column = mapping_config.get("target_column", template_column)
        else:
            source_column = template_column
            target_column = mapping_config

        if not source_column:
            raise MergeFolderError(f"文件 {file_path.name} 的字段映射缺少 source_column: {template_column}")

        try:
            col_idx = resolve_column_index_by_mode(
                target_column,
                headers,
                max_columns,
                "column_name",
                logger_instance=logger,
            )
        except ValueError as e:
            raise MergeFolderError(
                f"文件 {file_path.name} 无法解析字段 '{template_column}' 的目标列 '{target_column}': {e}"
            ) from e

        bindings.append((str(source_column), col_idx))

    return bindings


def _resolve_month_column(
    group_config: RuleGroupConfig | dict[str, Any],
    headers: dict[str, int],
    max_columns: int,
    writer: ExcelWriter | None,
    file_path: Path,
) -> int | None:
    del writer
    month_type_mapping = group_config.get("month_type_mapping", {})
    if not isinstance(month_type_mapping, dict) or not month_type_mapping.get("enabled"):
        return None

    target_column = month_type_mapping.get("target_column", "C")
    try:
        return resolve_column_index_by_mode(
            target_column,
            headers,
            max_columns,
            "column_name",
            logger_instance=logger,
        )
    except ValueError as e:
        raise MergeFolderError(
            f"文件 {file_path.name} 无法解析 month_type_mapping.target_column '{target_column}': {e}"
        ) from e


def _extract_headers(rows: list[list[Any]], header_row: int, file_path: Path) -> dict[str, int]:
    try:
        return extract_headers(rows, header_row)
    except ValueError as exc:
        raise MergeFolderError(f"文件 {file_path.name} 的 {exc}") from exc


def _read_all_rows(file_path: Path) -> list[list[Any]]:
    ext = file_path.suffix.lower()
    if ext == ".xlsx":
        return _read_xlsx_rows(file_path)
    if ext == ".xls":
        return _read_xls_rows(file_path)
    raise MergeFolderError(f"不支持的文件格式: {file_path.name}")


def _read_xlsx_rows(file_path: Path) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if worksheet is None:
            raise MergeFolderError(f"Excel 文件没有工作表: {file_path.name}")
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _read_xls_rows(file_path: Path) -> list[list[Any]]:
    workbook = xlrd.open_workbook(str(file_path))
    sheet = workbook.sheet_by_index(0)
    rows: list[list[Any]] = []
    for row_idx in range(sheet.nrows):
        row_values: list[Any] = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            row_values.append(_convert_xls_cell(cell, workbook.datemode))
        rows.append(row_values)
    return rows


def _convert_xls_cell(cell: Any, datemode: int) -> Any:
    return convert_xls_cell(cell, datemode)


def _get_cell_value(row_values: list[Any], column_index: int) -> Any:
    return get_cell_value(row_values, column_index)


def _is_empty_value(value: Any) -> bool:
    return is_empty_value(value)


def _infer_month_param_from_single_value(value: str, month_type_mapping: dict) -> str:
    normalized_value = str(value).strip()
    bonus_value = str(month_type_mapping.get("bonus_value", "年终奖"))
    compensation_value = str(month_type_mapping.get("compensation_value", "补偿金"))

    if normalized_value == bonus_value:
        return "年终奖"
    if normalized_value == compensation_value:
        return "补偿金"

    month_format = month_type_mapping.get("month_format", "{month}月收入")
    if not isinstance(month_format, str):
        raise MergeFolderError(f"month_format 必须是字符串，当前值: {month_format}")

    matched_months: list[str] = []
    for month in range(1, 13):
        month_token = f"{month:02d}"
        try:
            formatted = month_format.format(month=month_token)
        except KeyError as e:
            raise MergeFolderError(f"month_format 缺少变量: {e}") from e
        except Exception as e:
            raise MergeFolderError(f"month_format 格式错误: {e}") from e

        if formatted == normalized_value:
            matched_months.append(month_token)

    if len(matched_months) == 1:
        return matched_months[0]
    if len(matched_months) > 1:
        raise MergeFolderError(f"月类型值 '{normalized_value}' 对 month_format 匹配到多个月份: {matched_months}")
    raise MergeFolderError(f"无法从月类型值推断月份参数: '{normalized_value}'")


def _select_month_param_on_conflict(inferred_params: set[str]) -> str:
    """冲突时选择一个稳定的月份参数继续处理。"""
    numeric_months: list[str] = []
    text_params: list[str] = []

    for param in inferred_params:
        if param.isdigit():
            numeric_months.append(f"{int(param):02d}")
        else:
            text_params.append(param)

    if numeric_months:
        return sorted(set(numeric_months), key=lambda value: int(value))[0]
    if "年终奖" in text_params:
        return "年终奖"
    if "补偿金" in text_params:
        return "补偿金"
    return sorted(text_params)[0]


def _build_merge_output_group_config(
    group_config: RuleGroupConfig | dict[str, Any],
    keep_row_month_values: bool,
) -> RuleGroupConfig | dict[str, Any]:
    """
    构建用于合并输出的规则组配置。

    当 keep_row_month_values=True 且启用 month_type_mapping 时：
    - 关闭 month_type_mapping 的统一填充值行为
    - 通过 field_mappings 增加一条“原月类型值 -> 目标列”映射，按行回写原值
    """
    if not keep_row_month_values:
        return group_config

    month_type_mapping = group_config.get("month_type_mapping", {})
    if not isinstance(month_type_mapping, dict) or not month_type_mapping.get("enabled"):
        return group_config

    updated_config = dict(group_config)
    updated_field_mappings = dict(group_config.get("field_mappings", {}))
    target_column = month_type_mapping.get("target_column", "C")

    updated_field_mappings[MERGE_MONTH_SOURCE_COLUMN] = {
        "source_column": MERGE_MONTH_SOURCE_COLUMN,
        "target_column": target_column,
        "transform": "none",
    }
    updated_config["field_mappings"] = updated_field_mappings
    updated_config["month_type_mapping"] = {"enabled": False}

    return updated_config
