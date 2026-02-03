"""Excel文件读取器模块

支持读取 .xlsx, .csv, .xls 格式的文件，并将数据转换为字典列表。
"""

import logging
import csv
from pathlib import Path
from typing import List, Dict, Optional, Any

import openpyxl
import xlrd


logger = logging.getLogger(__name__)


class ExcelError(Exception):
    """Excel文件读取异常"""

    pass


class ExcelReader:
    """Excel文件读取器

    支持读取多种格式的Excel文件，并转换为字典列表。
    """

    def __init__(self, row_filter: Optional[Dict[str, Any]] = None):
        """初始化ExcelReader

        Args:
            row_filter: 行过滤配置，用于排除特定行
        """
        logger.debug("初始化ExcelReader")
        self.row_filter = row_filter or {}

    def read_excel(self, file_path: str) -> List[Dict[str, Any]]:
        """读取Excel文件并返回字典列表

        Args:
            file_path: Excel文件路径

        Returns:
            字典列表，每个字典代表一行数据

        Raises:
            FileNotFoundError: 文件不存在
            ExcelError: 文件格式无效或不支持
        """
        logger.info(f"开始读取文件: {file_path}")

        # 检查文件是否存在
        path = Path(file_path)
        if not path.exists():
            logger.error(f"文件不存在: {file_path}")
            raise FileNotFoundError(f"文件不存在: {file_path}")

        # 根据文件扩展名选择读取方式
        file_ext = path.suffix.lower()

        try:
            if file_ext == ".xlsx":
                return self._read_xlsx(file_path)
            elif file_ext == ".csv":
                return self._read_csv(file_path)
            elif file_ext == ".xls":
                return self._read_xls(file_path)
            else:
                logger.error(f"不支持的文件格式: {file_ext}")
                raise ExcelError(f"不支持的文件格式: {file_ext}")
        except ExcelError:
            # 重新抛出ExcelError
            raise
        except Exception as e:
            logger.error(f"读取文件失败: {file_path}, 错误: {e}")
            raise ExcelError(f"文件格式无效: {file_path}") from e

    def _is_empty_cell(self, value: Any) -> bool:
        """判断单元格是否为空"""
        if value is None:
            return True
        if isinstance(value, str):
            return not value.strip()
        return False

    def _convert_xls_cell(self, cell, datemode: int) -> Any:
        """将 .xls 单元格值转换为更合适的 Python 类型"""
        try:
            cell_type = cell.ctype
        except Exception:
            return cell.value

        # 空单元格
        empty_types = [getattr(xlrd, "XL_CELL_EMPTY", -1), getattr(xlrd, "XL_CELL_BLANK", -1)]
        if cell_type in empty_types:
            return None

        # 日期单元格
        if cell_type == getattr(xlrd, "XL_CELL_DATE", -1):
            try:
                return xlrd.xldate_as_datetime(cell.value, datemode)
            except Exception:
                return cell.value

        # 数字单元格
        if cell_type == getattr(xlrd, "XL_CELL_NUMBER", -1):
            try:
                if float(cell.value).is_integer():
                    return int(cell.value)
            except Exception:
                pass
            return cell.value

        # 布尔单元格
        if cell_type == getattr(xlrd, "XL_CELL_BOOLEAN", -1):
            return bool(cell.value)

        # 其他类型（文本、错误等）
        return cell.value

    def _should_skip_row(self, row_values: List[str], headers: Optional[List[str]]) -> bool:
        """检查是否应该跳过该行

        Args:
            row_values: 行数据（值列表）
            headers: 表头列表

        Returns:
            True: 应该跳过，False: 不跳过
        """
        if not self.row_filter:
            return False

        exclude_keywords = self.row_filter.get("exclude_keywords", [])
        if not exclude_keywords:
            return False

        # 将行转换为字典以便检查
        row_dict = {}
        if headers:
            for i, value in enumerate(row_values):
                if i < len(headers):
                    row_dict[headers[i]] = "" if value is None else str(value)

        # 检查是否包含排除关键字
        for keyword in exclude_keywords:
            if keyword in row_dict.values():
                logger.debug(f"检测到排除关键字 '{keyword}'，跳过该行")
                return True

        return False

    def _read_xlsx(self, file_path: str) -> List[Dict[str, Any]]:
        """读取.xlsx文件

        Args:
            file_path: .xlsx文件路径

        Returns:
            字典列表
        """
        logger.debug(f"使用openpyxl读取.xlsx文件: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheet = workbook.active  # 使用第一个工作表
            if sheet is None:
                raise ExcelError("Excel文件没有工作表")

            # 读取表头（第1行）
            headers = None
            data_rows = []

            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row_idx == 1:
                    # 第一行是表头
                    headers = [str(cell) if cell is not None else "" for cell in row]
                    logger.debug(f"提取表头: {headers}")
                else:
                    # 数据行
                    row_values = list(row)

                    # 检查是否为空行（所有单元格都为空）
                    if not row_values or all(self._is_empty_cell(cell) for cell in row_values):
                        logger.debug(f"跳过空行: 第{row_idx}行")
                        continue

                    # 应用行过滤（排除指定关键字）
                    if self._should_skip_row(row_values, headers):
                        logger.debug(f"跳过过滤行: 第{row_idx}行")
                        continue

                    # 将行转换为字典
                    if headers:
                        row_dict = {}
                        for col_idx, value in enumerate(row_values):
                            if col_idx < len(headers):
                                row_dict[headers[col_idx]] = value
                        data_rows.append(row_dict)

            workbook.close()
            logger.info(f"成功读取.xlsx文件，共 {len(data_rows)} 行数据")

            return data_rows

        except Exception as e:
            logger.error(f"读取.xlsx文件失败: {e}")
            raise ExcelError(f"无法读取.xlsx文件: {file_path}") from e

    def _read_csv(self, file_path: str) -> List[Dict[str, Any]]:
        """读取.csv文件

        Args:
            file_path: .csv文件路径

        Returns:
            字典列表
        """
        logger.debug(f"使用csv模块读取.csv文件: {file_path}")

        try:
            with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)

            if not rows:
                logger.warning("CSV文件为空")
                return []

            # 第一行是表头
            headers = rows[0]
            logger.debug(f"提取表头: {headers}")

            data_rows = []
            # 从第二行开始读取数据
            for row_idx, row in enumerate(rows[1:], start=2):
                # 检查是否为空行（所有单元格都为空）
                if not row or all(self._is_empty_cell(cell) for cell in row):
                    logger.debug(f"跳过空行: 第{row_idx}行")
                    continue

                # 应用行过滤（排除指定关键字）
                if self._should_skip_row(row, headers):
                    logger.debug(f"跳过过滤行: 第{row_idx}行")
                    continue

                # 将行转换为字典
                row_dict = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = value
                data_rows.append(row_dict)

            logger.info(f"成功读取.csv文件，共 {len(data_rows)} 行数据")

            return data_rows

        except Exception as e:
            logger.error(f"读取.csv文件失败: {e}")
            raise ExcelError(f"无法读取.csv文件: {file_path}") from e

    def _read_xls(self, file_path: str) -> List[Dict[str, Any]]:
        """读取.xls文件

        Args:
            file_path: .xls文件路径

        Returns:
            字典列表
        """
        logger.debug(f"使用xlrd读取.xls文件: {file_path}")

        try:
            workbook = xlrd.open_workbook(file_path)
            sheet = workbook.sheet_by_index(0)  # 使用第一个工作表

            # 读取表头（第1行）
            headers = []
            for col_idx in range(sheet.ncols):
                cell_value = sheet.cell_value(0, col_idx)
                headers.append(str(cell_value) if cell_value is not None else "")
            logger.debug(f"提取表头: {headers}")

            data_rows = []
            # 从第2行开始读取数据
            for row_idx in range(1, sheet.nrows):
                row_values = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    cell_value = self._convert_xls_cell(cell, workbook.datemode)
                    row_values.append(cell_value)

                # 检查是否为空行（所有单元格都为空）
                if not row_values or all(self._is_empty_cell(cell) for cell in row_values):
                    logger.debug(f"跳过空行: 第{row_idx + 1}行")
                    continue

                # 应用行过滤（排除指定关键字）
                if self._should_skip_row(row_values, headers):
                    logger.debug(f"跳过过滤行: 第{row_idx + 1}行")
                    continue

                # 将行转换为字典
                row_dict = {}
                for col_idx, value in enumerate(row_values):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = value
                data_rows.append(row_dict)

            logger.info(f"成功读取.xls文件，共 {len(data_rows)} 行数据")

            return data_rows

        except Exception as e:
            logger.error(f"读取.xls文件失败: {e}")
            raise ExcelError(f"无法读取.xls文件: {file_path}") from e
