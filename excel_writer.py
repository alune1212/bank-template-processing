"""Excel写入器模块

提供Excel、CSV、XLS文件写入功能，支持保留格式、字段映射、固定值、自动编号等高级功能。
"""

import csv
import logging
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    openpyxl = None
    InvalidFileException = None

try:
    import xlrd
except ImportError:
    xlrd = None

try:
    import xlwt
except ImportError:
    xlwt = None

from config_loader import ConfigError


logger = logging.getLogger(__name__)


class ExcelError(Exception):
    """Excel操作失败异常"""

    pass


class ExcelWriter:
    """Excel写入器，支持.xlsx、.csv、.xls三种格式"""

    def __init__(self):
        """初始化写入器"""
        logger.debug("ExcelWriter 初始化")

    def write_excel(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> None:
        """
        将数据写入Excel文件（支持.xlsx、.csv、.xls三种格式）

        Args:
            template_path: 模板文件路径
            data: 要写入的数据列表（字典列表）
            field_mappings: 字段映射配置
            output_path: 输出文件路径
            header_row: 表头行号（从1开始）
            start_row: 数据起始行号（从1开始）
            mapping_mode: 映射模式（column_name优先列名，column_index优先列索引）
            fixed_values: 固定值配置（例如：{"A": "固定文本"}）
            auto_number: 自动编号配置（例如：{"enabled": True, "column": "A", "start_from": 1}）
            bank_branch_mapping: 银行支行映射（例如：{"enabled": True, "source_column": "支行", "target_column": "B"}）
            month_type_mapping: 月类型映射（例如：{"enabled": True, "target_column": "C", "bonus_value": "年终奖", "compensation_value": "补偿金"}）
            month_param: 月参数（例如："1"或"01"或"年终奖"或"补偿金"）

        Raises:
            ConfigError: 配置无效时抛出
            ExcelError: 文件操作失败时抛出
        """
        logger.info(f"开始写入Excel文件: {template_path} -> {output_path}")
        logger.info(f"数据行数: {len(data)}")

        # 验证配置
        if start_row <= header_row:
            error_msg = (
                f"配置错误: start_row ({start_row}) 必须大于 header_row ({header_row})"
            )
            logger.error(error_msg)
            raise ConfigError(error_msg)

        # 根据文件扩展名选择写入方式
        ext = Path(template_path).suffix.lower()

        try:
            if ext == ".xlsx":
                self._write_xlsx(
                    template_path,
                    data,
                    field_mappings,
                    output_path,
                    header_row,
                    start_row,
                    mapping_mode,
                    fixed_values,
                    auto_number,
                    bank_branch_mapping,
                    month_type_mapping,
                    month_param,
                )
            elif ext == ".csv":
                self._write_csv(
                    template_path,
                    data,
                    field_mappings,
                    output_path,
                    header_row,
                    start_row,
                    mapping_mode,
                    fixed_values,
                    auto_number,
                    bank_branch_mapping,
                    month_type_mapping,
                    month_param,
                )
            elif ext == ".xls":
                self._write_xls(
                    template_path,
                    data,
                    field_mappings,
                    output_path,
                    header_row,
                    start_row,
                    mapping_mode,
                    fixed_values,
                    auto_number,
                    bank_branch_mapping,
                    month_type_mapping,
                    month_param,
                )
            else:
                error_msg = f"不支持的文件格式: {ext}"
                logger.error(error_msg)
                raise ExcelError(error_msg)

        except Exception as e:
            if isinstance(e, (ConfigError, ExcelError, FileNotFoundError)):
                raise
            error_msg = f"写入文件失败: {e}"
            logger.error(error_msg, exc_info=True)
            raise ExcelError(error_msg) from e

        logger.info(f"文件写入成功: {output_path}")

    def _write_xlsx(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> None:
        """使用openpyxl写入.xlsx文件，保留格式、公式、合并单元格"""
        logger.debug(f"使用openpyxl写入xlsx文件: {template_path}")

        if openpyxl is None:
            raise ExcelError("openpyxl未安装，无法处理.xlsx文件")

        # 加载工作簿
        try:
            wb = openpyxl.load_workbook(template_path)
        except InvalidFileException as e:
            raise ExcelError(f"无效的Excel文件: {e}") from e

        # 使用第一个工作表
        ws = wb.active

        # 读取表头（用于字段映射）
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(header_row, col_idx).value
            if cell_value:
                headers[str(cell_value).strip()] = col_idx

        logger.debug(f"读取到 {len(headers)} 个表头字段")

        # 清除从start_row开始的所有行（保留header_row）
        logger.debug(f"清除从第 {start_row} 行开始的数据")
        max_row = ws.max_row
        for row_idx in range(max_row, start_row - 1, -1):
            ws.delete_rows(row_idx)

        # 应用字段映射并写入数据
        self._write_data_to_worksheet(
            ws,
            data,
            field_mappings,
            headers,
            start_row,
            mapping_mode,
            fixed_values,
            auto_number,
            bank_branch_mapping,
            month_type_mapping,
            month_param,
        )

        # 保存文件
        wb.save(output_path)
        logger.debug(f"xlsx文件已保存: {output_path}")

    def _write_csv(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> None:
        """使用csv模块写入.csv文件（UTF-8编码，无格式保留）"""
        logger.debug(f"使用csv模块写入csv文件: {template_path}")

        # 读取模板文件
        with open(template_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        if len(rows) < header_row:
            raise ExcelError(f"模板文件行数不足，无法读取表头行: {header_row}")

        # 读取表头
        headers = {}
        header_row_data = rows[header_row - 1]  # 转换为0-index
        for col_idx, cell_value in enumerate(header_row_data, 1):  # 1-index
            if cell_value:
                headers[cell_value.strip()] = col_idx

        logger.debug(f"读取到 {len(headers)} 个表头字段")

        # 保留header_row行之前的所有内容
        output_rows = rows[:header_row]

        # 处理数据并转换为行列表
        data_rows = self._process_data_to_rows(
            data,
            field_mappings,
            headers,
            len(header_row_data),
            mapping_mode,
            fixed_values,
            auto_number,
            bank_branch_mapping,
            month_type_mapping,
            month_param,
        )

        # 添加数据行
        output_rows.extend(data_rows)

        # 写入输出文件
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(output_rows)

        logger.debug(f"csv文件已保存: {output_path}")

    def _write_xls(
        self,
        template_path: str,
        data: list,
        field_mappings: dict,
        output_path: str,
        header_row: int,
        start_row: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> None:
        """使用xlwt写入.xls文件（基本格式，无公式保留）"""
        logger.debug(f"使用xlwt写入xls文件: {template_path}")

        if xlwt is None:
            raise ExcelError("xlwt未安装，无法处理.xls文件")

        if xlrd is None:
            raise ExcelError("xlrd未安装，无法读取.xls模板文件")

        try:
            wb_template = xlrd.open_workbook(template_path, formatting_info=False)
        except Exception as e:
            raise ExcelError(f"无法读取Excel文件: {e}") from e

        ws_template = wb_template.sheet_by_index(0)

        headers = {}
        for col_idx in range(ws_template.ncols):
            cell_value = ws_template.cell_value(
                header_row - 1, col_idx
            )  # xlrd使用0-index
            if cell_value:
                headers[str(cell_value).strip()] = col_idx + 1

        logger.debug(f"读取到 {len(headers)} 个表头字段")

        wb_output = xlwt.Workbook(encoding="utf-8")
        ws_output = wb_output.add_sheet("Sheet1")

        for row_idx in range(header_row):
            for col_idx in range(ws_template.ncols):
                cell_value = ws_template.cell_value(row_idx, col_idx)
                if cell_value:
                    ws_output.write(row_idx, col_idx, cell_value)

        self._write_data_to_xls_sheet(
            ws_output,
            data,
            field_mappings,
            headers,
            start_row,
            ws_template.ncols,
            mapping_mode,
            fixed_values,
            auto_number,
            bank_branch_mapping,
            month_type_mapping,
            month_param,
        )

        wb_output.save(output_path)
        logger.debug(f"xls文件已保存: {output_path}")

    def _process_data_to_rows(
        self,
        data: list,
        field_mappings: dict,
        headers: dict,
        max_columns: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> list:
        """
        处理数据并转换为行列表（用于CSV写入）

        Returns:
            list of lists: 数据行列表
        """
        result_rows = []

        current_number = None
        if auto_number and auto_number.get("enabled"):
            current_number = auto_number.get("start_from", 1)

        month_value = None
        if month_type_mapping and month_type_mapping.get("enabled"):
            month_value = self._calculate_month_value(month_param, month_type_mapping)

        for row_idx, row_data in enumerate(data):
            row_output = [""] * max_columns

            # 应用字段映射
            for template_column, mapping_config in field_mappings.items():
                # 兼容两种配置格式：
                # 1. 新格式：{ "模板列名": { "source_column": "输入列名", ... } }
                # 2. 旧格式：{ "输入列名": "模板列名" }
                if isinstance(mapping_config, dict):
                    # 新格式
                    source_column = mapping_config.get("source_column")
                    target_column = mapping_config.get("target_column", template_column)
                else:
                    # 旧格式：key是输入列名，value是模板列名
                    source_column = template_column
                    target_column = mapping_config

                # 获取源数据值
                value = row_data.get(source_column, "")

                # 获取目标列索引
                if mapping_mode == "column_name":
                    # 使用列名
                    if target_column in headers:
                        col_idx = headers[target_column]
                    else:
                        continue
                else:
                    # 使用列索引（支持"A"这样的Excel列标识）
                    col_idx = self._column_letter_to_index(str(target_column).upper())

                # 写入数据（转换为0-index）
                if 1 <= col_idx <= max_columns:
                    row_output[col_idx - 1] = str(value) if value is not None else ""

            # 应用固定值
            if fixed_values:
                for column, value in fixed_values.items():
                    col_idx = self._column_letter_to_index(column)
                    if 1 <= col_idx <= max_columns:
                        row_output[col_idx - 1] = str(value)

            if auto_number and auto_number.get("enabled"):
                column = auto_number.get("column", "A")
                col_idx = self._column_letter_to_index(column)
                if 1 <= col_idx <= max_columns:
                    row_output[col_idx - 1] = str(current_number)
                if current_number is not None:
                    current_number += 1

            # 应用银行支行映射
            if bank_branch_mapping and bank_branch_mapping.get("enabled"):
                source_column = bank_branch_mapping.get("source_column")
                target_column = bank_branch_mapping.get("target_column", "B")
                branch_info = row_data.get(source_column, "")
                if branch_info:
                    col_idx = self._column_letter_to_index(target_column)
                    if 1 <= col_idx <= max_columns:
                        row_output[col_idx - 1] = str(branch_info)

            if month_value is not None and month_type_mapping:
                target_column = month_type_mapping.get("target_column", "C")
                col_idx = self._column_letter_to_index(target_column)
                if 1 <= col_idx <= max_columns:
                    row_output[col_idx - 1] = str(month_value)

            result_rows.append(row_output)

        return result_rows

    def _write_data_to_worksheet(
        self,
        ws,
        data: list,
        field_mappings: dict,
        headers: dict,
        start_row: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> None:
        """将数据写入openpyxl工作表"""
        current_number = None
        if auto_number and auto_number.get("enabled"):
            current_number = auto_number.get("start_from", 1)

        month_value = None
        if month_type_mapping and month_type_mapping.get("enabled"):
            month_value = self._calculate_month_value(month_param, month_type_mapping)

        for row_idx, row_data in enumerate(data):
            output_row_idx = start_row + row_idx

            # 应用字段映射
            for template_column, mapping_config in field_mappings.items():
                # 兼容两种配置格式：
                # 1. 新格式：{ "模板列名": { "source_column": "输入列名", ... } }
                # 2. 旧格式：{ "输入列名": "模板列名" }
                if isinstance(mapping_config, dict):
                    # 新格式
                    source_column = mapping_config.get("source_column")
                    target_column = mapping_config.get("target_column", template_column)
                else:
                    # 旧格式：key是输入列名，value是模板列名
                    source_column = template_column
                    target_column = mapping_config

                # 获取源数据值
                value = row_data.get(source_column, "")

                # 获取目标列索引
                if mapping_mode == "column_name":
                    # 使用列名
                    if target_column in headers:
                        col_idx = headers[target_column]
                    else:
                        continue
                else:
                    # 使用列索引（支持"A"这样的Excel列标识）
                    col_idx = self._column_letter_to_index(str(target_column).upper())

                # 写入数据
                ws.cell(output_row_idx, col_idx, value)

            # 应用固定值
            if fixed_values:
                for column, value in fixed_values.items():
                    col_idx = self._column_letter_to_index(column)
                    ws.cell(output_row_idx, col_idx, value)

            if auto_number and auto_number.get("enabled"):
                column = auto_number.get("column", "A")
                col_idx = self._column_letter_to_index(column)
                ws.cell(output_row_idx, col_idx, current_number)
                if current_number is not None:
                    current_number += 1

            if bank_branch_mapping and bank_branch_mapping.get("enabled"):
                source_column = bank_branch_mapping.get("source_column")
                target_column = bank_branch_mapping.get("target_column", "B")
                branch_info = row_data.get(source_column, "")
                if branch_info:
                    col_idx = self._column_letter_to_index(target_column)
                    ws.cell(output_row_idx, col_idx, branch_info)

            if month_value is not None and month_type_mapping:
                target_column = month_type_mapping.get("target_column", "C")
                col_idx = self._column_letter_to_index(target_column)
                ws.cell(output_row_idx, col_idx, month_value)

        logger.debug(f"已写入 {len(data)} 行数据到工作表")

    def _write_data_to_xls_sheet(
        self,
        ws,
        data: list,
        field_mappings: dict,
        headers: dict,
        start_row: int,
        max_columns: int,
        mapping_mode: str,
        fixed_values: Optional[dict] = None,
        auto_number: Optional[dict] = None,
        bank_branch_mapping: Optional[dict] = None,
        month_type_mapping: Optional[dict] = None,
        month_param: Optional[str] = None,
    ) -> None:
        """将数据写入xlwt工作表（使用0-index）"""
        current_number = None
        if auto_number and auto_number.get("enabled"):
            current_number = auto_number.get("start_from", 1)

        month_value = None
        if month_type_mapping and month_type_mapping.get("enabled"):
            month_value = self._calculate_month_value(month_param, month_type_mapping)

        for row_idx, row_data in enumerate(data):
            output_row_idx = start_row + row_idx

            # 应用字段映射
            for template_column, mapping_config in field_mappings.items():
                # 兼容两种配置格式：
                # 1. 新格式：{ "模板列名": { "source_column": "输入列名", ... } }
                # 2. 旧格式：{ "输入列名": "模板列名" }
                if isinstance(mapping_config, dict):
                    # 新格式
                    source_column = mapping_config.get("source_column")
                    target_column = mapping_config.get("target_column", template_column)
                else:
                    # 旧格式：key是输入列名，value是模板列名
                    source_column = template_column
                    target_column = mapping_config

                # 获取源数据值
                value = row_data.get(source_column, "")

                # 获取目标列索引
                if mapping_mode == "column_name":
                    # 使用列名
                    if target_column in headers:
                        col_idx = headers[target_column]
                    else:
                        continue
                else:
                    # 使用列索引（支持"A"这样的Excel列标识）
                    col_idx = self._column_letter_to_index(str(target_column).upper())

                # 写入数据（转换为0-index）
                if 1 <= col_idx <= max_columns:
                    ws.write(output_row_idx - 1, col_idx - 1, value)

            # 应用固定值
            if fixed_values:
                for column, value in fixed_values.items():
                    col_idx = self._column_letter_to_index(column)
                    if 1 <= col_idx <= max_columns:
                        ws.write(output_row_idx - 1, col_idx - 1, value)

            if auto_number and auto_number.get("enabled"):
                column = auto_number.get("column", "A")
                col_idx = self._column_letter_to_index(column)
                if 1 <= col_idx <= max_columns:
                    ws.write(output_row_idx - 1, col_idx - 1, current_number)
                if current_number is not None:
                    current_number += 1

            if bank_branch_mapping and bank_branch_mapping.get("enabled"):
                source_column = bank_branch_mapping.get("source_column")
                target_column = bank_branch_mapping.get("target_column", "B")
                branch_info = row_data.get(source_column, "")
                if branch_info:
                    col_idx = self._column_letter_to_index(target_column)
                    if 1 <= col_idx <= max_columns:
                        ws.write(output_row_idx - 1, col_idx - 1, branch_info)

            if month_value is not None and month_type_mapping:
                target_column = month_type_mapping.get("target_column", "C")
                col_idx = self._column_letter_to_index(target_column)
                if 1 <= col_idx <= max_columns:
                    ws.write(output_row_idx - 1, col_idx - 1, month_value)

        logger.debug(f"已写入 {len(data)} 行数据到xls工作表")

    def _calculate_month_value(
        self, month_param: Optional[str], month_type_mapping: dict
    ) -> Optional[str]:
        """
        计算month类型映射的值

        Args:
            month_param: 月参数（例如："1"或"01"或"年终奖"或"补偿金"）
            month_type_mapping: 月类型映射配置

        Returns:
            str: 计算后的值
        """
        if month_param is None:
            return None

        # 检查是否为月份数字
        try:
            month_num = int(month_param)
            if 1 <= month_num <= 12:
                # 格式化为"XX月收入"
                return f"{month_num:02d}月收入"
        except ValueError:
            pass

        # 检查是否为"年终奖"
        if month_param == "年终奖":
            return month_type_mapping.get("bonus_value", "年终奖")

        # 检查是否为"补偿金"
        if month_param == "补偿金":
            return month_type_mapping.get("compensation_value", "补偿金")

        return None

    def _column_letter_to_index(self, column: str) -> int:
        """
        将Excel列标识转换为索引（A=1, B=2, ..., AA=27）

        Args:
            column: 列标识（例如："A", "B", "AA"）

        Returns:
            int: 列索引（从1开始）
        """
        column = column.upper()
        index = 0
        for char in column:
            index = index * 26 + (ord(char) - ord("A") + 1)
        return index
